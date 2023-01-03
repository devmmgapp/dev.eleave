from datetime import date, datetime, timedelta 
from flask import jsonify, request, current_app, send_file, Blueprint
from flask import session, request, jsonify
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from io import BytesIO 
from bson.objectid import ObjectId
from datetime import datetime
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.styles import borders
from openpyxl.styles.borders import Border
from openpyxl.styles.alignment import Alignment
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from copy import copy

import pandas as pd
import gridfs
import calendar
import smtplib
import socket
import json
import os
import re
from dotenv import load_dotenv
load_dotenv()

import checkLogged
from my_app import db,  client

eleave = Blueprint('eleave', __name__)

#########################################################################################################
## e-leave
#########################################################################################################

#db
eleaveDtl = db["eleave_dtl"]
holidays = db["holidays"]
leaveTypes = db["leave_types"]
leaveGroups = db["leave_groups"]
maintenance = db["eleave_maintenance"]
reportMap = db["fileDirectory"]

#Global Constant
status = list(maintenance.find({"table": { "$eq" : "globalConstant"}}))
df = pd.DataFrame(status)

## parameters
#leaveOffice = "HKG"
#leaveYear = 2022
#leaveRacf = 'NF1KWY'
#leaveType = "LVE02"
#leaveType = "Casual Leave"
#approver = "NF1VCC"
#leaveType = "Work From Home"
#leaveApplying = [{"startDate": "2022-07-19", "startTime": "AM", "endDate": "2022-07-19", "endTime": "PM"},
#                 {"startDate" : "2022-07-20", "startTime": "PM", "endDate": "2022-07-20", "endTime": "PM"},
#                 { "startDate" : "2022-08-01", "startTime": "AM", "endDate" : "2022-08-02", "endTime": "AM"}
#                ]
#leaveApplyingScreen =  [   {"startDate": "2022-07-19", "startTime": "Full Day", "endDate": "2022-07-19", "endTime": "Full Day"},
#                            {"startDate": "2022-07-20", "startTime": "Half Day - PM", "endDate": "2022-07-20", "endTime": "Half Day - PM"},
#                            {"startDate": "2022-08-01", "startTime": "Full Day", "endDate": "2022-08-02", "endTime": "Half Day - AM"}
#                        ]
#leaveApplying = [{"startDate": "2022-07-04", "startTime": "AM", "endDate": "2022-07-04", "endTime": "AM"}]
#leaveApplyingScreen = [{ "startDate": "2022-07-04", "startTime": "Half Day - AM", "endDate": "2022-07-15", "endTime": "Half Day - AM"}]


leaveTypeLst = []
leaveGroupLst = []


@eleave.route('/api/getUserList',methods=['POST'])
@checkLogged.check_logged
def getUserList():            
       
    ## loading staff list  
    query = { "staff":1}        
    query_filter =  {"staff.status": { "$eq" : "ACTIVE"}}                 
    userList = []     
    col = eleaveDtl 
    results = col.find(query_filter, query)    ## eleave_dtl
    for result in results:   
        userList.append({ 'racf': result['staff']['racf'], 'name' : result['staff']['name']})

    try:
        if len(userList) > 0:                             
            return  jsonify({'userList' : userList}), 200   
        else:
            return  jsonify({'error_message' : 'Error to get user list.  Please contact regional PBT !'}), 501
    except:
        return jsonify({'error_message' : 'Error to get user list.  Please contact regional PBT !'}), 501         





## Functions 

def genReport(psWS, psRptDict, psRptFormat):
    for lstKey, lstValue in psRptDict.items():
        try:
            if lstKey in psRptFormat["cell"]:
                if (isinstance(lstValue, list)) == False:
                    psWS[(psRptFormat["cell"][lstKey])].value = lstValue
                else:
                    row = 0
                    col = 0
                    for lveRow in lstValue:
                        for key, v in lveRow.items():
                            if psRptFormat["cell"][lstKey]["next_record"] == "Row":
                                mcell = True
                                while mcell == True:
                                    colId = column_index_from_string(coordinate_from_string(psRptFormat["cell"][lstKey]["start_cell"])[0]) + col
                                    rowId = coordinate_from_string(psRptFormat["cell"][lstKey]["start_cell"])[1] + row
                                    if not isinstance(psWS.cell(row=rowId, column=colId), MergedCell):
                                        mcell = False
                                    else:
                                        mcell = True
                                        col += 1                                                                                       
                                psWS.cell(row=rowId, column=colId, value=v)
                                psWS.cell(row=rowId, column=colId).border = Border(left=borders.Side(border_style='thin', color="FF000000", style=None), 
                                right=borders.Side(border_style='thin', color="FF000000", style=None), 
                                top=borders.Side(border_style='thin', color="FF000000", style=None),
                                bottom=borders.Side(border_style='thin', color="FF000000", style=None))
                                col += 1 
                        row += 1
                        col = 0


        except:
            print ("Error found")
            return "Error", 701
    return "OK", 0    


def genApplyForm(ws, approvalRecordLst, rpt):

    # Basic information
    for excel_range in rpt['cell']:
        #Static information not included the office display and leave entry
        if excel_range != "LeaveDetail"  and excel_range != "applicantName":
            col_index = column_index_from_string(coordinate_from_string(rpt['cell'][excel_range])[0])
            row_index = (coordinate_from_string(rpt['cell'][excel_range])[1])
            ws.cell(row=row_index, column=col_index, value=approvalRecordLst[0][excel_range])

    # Count number of record
    record_count = len(approvalRecordLst[0]['details'])

    # Handle excel_range = "LeaveDetail"
    for rows in range ((coordinate_from_string(rpt['cell']['LeaveDetail'])[1]), record_count*2 +(coordinate_from_string(rpt['cell']['LeaveDetail'])[1]), 2):
        if rows == 36:
            col_index = column_index_from_string(coordinate_from_string(rpt['cell']['LeaveDetail'])[0])
            row_index = (coordinate_from_string(rpt['cell']['LeaveDetail'])[1])
            # Leave Start Date
            ws.cell(row=row_index, column=col_index, value=getMMDDYYYY(approvalRecordLst[0]['details'][rows-36]['startDate']))
            # Workday Name of start date
            ws.cell(row=row_index, column=col_index+ 3, value= "(" + getWorkdayName(approvalRecordLst[0]['details'][rows-36]['startDate']) + ")")
            # Full Day / AM / PM
            ws.cell(row=row_index, column=col_index+ 6, value=(approvalRecordLst[0]['details'][rows-36]['startTime']))
            # Leave End Date
            ws.cell(row=row_index, column=col_index+ 12, value=getMMDDYYYY(approvalRecordLst[0]['details'][rows-36]['endDate']))
            # Workday Name of end date
            ws.cell(row=row_index, column=col_index+ 15, value= "(" + getWorkdayName(approvalRecordLst[0]['details'][rows-36]['endDate']) + ")")
            # Full Day / AM / PM
            ws.cell(row=row_index, column=col_index+ 18, value=(approvalRecordLst[0]['details'][rows-36]['endTime']))
            # No of Working Days
            ws.cell(row=row_index, column=col_index+ 22, value=(approvalRecordLst[0]['details'][rows-36]['workday']))
            # No of Calendar Day
            ws.cell(row=row_index, column=col_index+ 28, value=(approvalRecordLst[0]['details'][rows-36]['calendarDay']))

        elif rows > 36:
            ws.insert_rows(rows, 2)
            # row and column index to output file
            
            index = int((rows-36)/2)
            col_index = column_index_from_string(coordinate_from_string(rpt['cell']['LeaveDetail'])[0])

            # Formatting, format must be the same as the first row of leave detail
            for n in range (0, ws.max_column):
                ws.cell(row=rows, column= 1 + n).value = copy(ws.cell(row = rows - 2, column= 1 + n).value)
                ws.cell(row=rows + 1, column= 1 + n).value = copy(ws.cell(row = rows - 1, column= 1 + n).value)
                ws.cell(row=rows, column= 1 + n).fill = copy(ws.cell(row = rows - 2, column= 1 + n).fill)
                ws.cell(row=rows + 1, column= 1 + n).fill = copy(ws.cell(row = rows - 1, column= 1 + n).fill)                
                ws.cell(row=rows, column= 1 + n).font = copy(ws.cell(row = rows - 2, column= 1 + n).font)
                ws.cell(row=rows + 1, column= 1 + n).font = copy(ws.cell(row = rows - 1, column= 1 + n).font)    
                ws.cell(row=rows, column= 1 + n).number_format = copy(ws.cell(row = rows - 2, column= 1 + n).number_format)
                ws.cell(row=rows + 1, column= 1 + n).number_format = copy(ws.cell(row = rows - 1, column= 1 + n).number_format)    
                ws.cell(row=rows, column= 1 + n).border = copy(ws.cell(row = rows - 2, column= 1 + n).border)
                ws.cell(row=rows + 1, column= 1 + n).border = copy(ws.cell(row = rows - 1, column= 1 + n).border)
                ws.cell(row=rows, column= 1 + n).alignment = copy(ws.cell(row = rows - 2, column= 1 + n).alignment)
                ws.cell(row=rows + 1, column= 1 + n).alignment = copy(ws.cell(row = rows - 1, column= 1 + n).alignment)
            
            # Merge Cells
            ws.merge_cells(start_row=rows, start_column=col_index, end_row=rows, end_column=col_index + 2 )
            ws.merge_cells(start_row=rows, start_column=col_index + 6, end_row=rows, end_column=col_index + 7)
            ws.merge_cells(start_row=rows, start_column=col_index + 12, end_row=rows, end_column=col_index + 14)
            ws.merge_cells(start_row=rows, start_column=col_index + 15, end_row=rows, end_column=col_index + 17)
            ws.merge_cells(start_row=rows, start_column=col_index + 18, end_row=rows, end_column=col_index + 19)
            ws.merge_cells(start_row=rows, start_column=col_index + 22, end_row=rows, end_column=col_index + 25)
            ws.merge_cells(start_row=rows, start_column=col_index + 28, end_row=rows, end_column=col_index + 32)

            # Leave Start Date
            ws.cell(row=rows, column=col_index, value=getMMDDYYYY(approvalRecordLst[0]['details'][index]['startDate']))
            # Workday Name of start date
            ws.cell(row=rows, column=col_index+ 3, value= "(" + getWorkdayName(approvalRecordLst[0]['details'][index]['startDate']) + ")")
            # Full Day / AM / PM
            ws.cell(row=rows, column=col_index+ 6, value=(approvalRecordLst[0]['details'][index]['startTime']))
            # Leave End Date
            ws.cell(row=rows, column=col_index+ 12, value=getMMDDYYYY(approvalRecordLst[0]['details'][index]['endDate']))
            # Workday Name of end date
            ws.cell(row=rows, column=col_index+ 15, value= "(" + getWorkdayName(approvalRecordLst[0]['details'][index]['endDate']) + ")")
            # Full Day / AM / PM
            ws.cell(row=rows, column=col_index+ 18, value=(approvalRecordLst[0]['details'][index]['endTime']))
            # No of Working Days
            ws.cell(row=rows, column=col_index+ 22, value=(approvalRecordLst[0]['details'][index]['workday']))
            # No of Calendar Day
            ws.cell(row=rows, column=col_index+ 28, value=(approvalRecordLst[0]['details'][index]['calendarDay']))
    
    # Handle excel_range = "applicantName"
    row_applicant = record_count*2 +(coordinate_from_string(rpt['cell']['LeaveDetail'])[1]) + 1
    col_index = column_index_from_string(coordinate_from_string(rpt['cell']["applicantName"])[0])
    row_index = row_applicant
    ws.cell(row=row_index, column=col_index, value=(approvalRecordLst[0]['staff']))
    ws.cell(row=row_index+1, column=col_index, value=(approvalRecordLst[0]['position']))
    ws.cell(row=row_index, column=col_index+19, value=(approvalRecordLst[0]['submit_date'])).alignment = Alignment(horizontal='center', vertical = 'center', wrap_text=True, wrapText=True)
    ws.merge_cells(start_row=row_index, start_column=col_index+19, end_row=row_index, end_column=col_index + 25)

    # Approver list below the Applicant Name in the form
    row_index = row_applicant + 2
    
    # Count number of record
    record_count = 1
    if len(approvalRecordLst[0]['approver2']) > 0: record_count = record_count + 1
    if len(approvalRecordLst[0]['approver3']) > 0: record_count = record_count + 1
    if record_count == 2: ws.insert_rows(row_index + 2, 2)
    if record_count == 3: ws.insert_rows(row_index + 2, 4)

    for n in range (0, ws.max_column):
        for i in range (1, record_count):
            ws.cell(row=row_index + i * 2, column= 1 + n).value = copy(ws.cell(row = (row_index + i * 2) - 2, column= 1 + n).value)
            ws.cell(row=(row_index + i * 2) + 1, column= 1 + n).value = copy(ws.cell(row = (row_index + i * 2) - 1, column= 1 + n).value)
            if n == 0 and i == 1:
                ws.cell(row=row_index + i * 2, column= 1 + n).value = str(ws.cell(row=row_index + i * 2, column= 1 + n).value).replace("1st","2nd")
            elif n == 0 and i == 2:
                ws.cell(row=row_index + i * 2, column= 1 + n).value = str(ws.cell(row=row_index + i * 2, column= 1 + n).value).replace("2nd","3rd")
            ws.cell(row=row_index + i * 2, column= 1 + n).fill = copy(ws.cell(row = (row_index + i * 2) - 2, column= 1 + n).fill)
            ws.cell(row=(row_index + i * 2) + 1, column= 1 + n).fill = copy(ws.cell(row = (row_index + i * 2) - 1, column= 1 + n).fill)                
            ws.cell(row=row_index + i * 2, column= 1 + n).font = copy(ws.cell(row = (row_index + i * 2) - 2, column= 1 + n).font)
            ws.cell(row=(row_index + i * 2) + 1, column= 1 + n).font = copy(ws.cell(row = (row_index + i * 2) - 1, column= 1 + n).font)     
            ws.cell(row=row_index + i * 2, column= 1 + n).number_format = copy(ws.cell(row = (row_index + i * 2) - 2, column= 1 + n).number_format)
            ws.cell(row=(row_index + i * 2) + 1, column= 1 + n).number_format = copy(ws.cell(row = (row_index + i * 2) - 1, column= 1 + n).number_format)  
            ws.cell(row=row_index + i * 2, column= 1 + n).border = copy(ws.cell(row = (row_index + i * 2) - 2, column= 1 + n).border)
            ws.cell(row=(row_index + i * 2) + 1, column= 1 + n).border = copy(ws.cell(row = (row_index + i * 2) - 1, column= 1 + n).border)
            ws.cell(row=row_index + i * 2, column= 1 + n).alignment = copy(ws.cell(row = (row_index + i * 2) - 2, column= 1 + n).alignment)
            ws.cell(row=(row_index + i * 2) + 1, column= 1 + n).alignment = copy(ws.cell(row = (row_index + i * 2) - 1, column= 1 + n).alignment)

    #approver list out
    for i in range (0, record_count):
        ws.cell(row=row_index + i * 2, column=col_index, value=(approvalRecordLst[0]['approver'+str(i+1)]))
        ws.cell(row=(row_index + i * 2) + 1, column=col_index, value=(approvalRecordLst[0]['approver_pos'+str(i+1)]))
        if len(str((approvalRecordLst[0]['approval_date'+str(i+1)]))) > 0:
            ws.cell(row=row_index + i * 2, column=col_index+19, value= "APPROVED \n" + (approvalRecordLst[0]['approval_date'+str(i+1)])).alignment = Alignment(horizontal='center', vertical = 'center', wrap_text=True, wrapText=True)
        ws.merge_cells(start_row=row_index + i * 2, start_column=col_index + 19, end_row=row_index + i * 2, end_column=col_index + 25)   


#convert date from string (yyyy-mm-dd) to date format.
#parameter : must be in string (yyyy-mm-dd) format
#return :
#Date in datetime format.
def str2Date (psDateStr):
    return datetime.strptime(psDateStr, "%Y-%m-%d")

#covert date to string 
#parameter : must be in datetime format
#return :
#date in string format : "YYYY-MM-DD"
def date2Str(psDate):
    return datetime.strftime(psDate, "%Y-%m-%d")

# get staff record 
# parameter:
# psRacf - RACF of the user
# return:
# staff record in MongoDB.
def getStaffRecord (psRacf):
    staffRecord = eleaveDtl.find_one ( {"staff.racf" : { '$regex' : psRacf, '$options' : "i"} , "staff.status": { '$regex': "ACTIVE", '$options': "i"} } )
    return(staffRecord)

def getLeaveTypes():
    global leaveTypeLst
    leaveTypeLst = list(leaveTypes.find({}))

def getLeaveGroups():
    global leaveGroupLst
    leaveGroupLst = list(leaveGroups.find({}))

# get long ref no. for displaying 
# parameter:
# psOffice - hr_office of the staff
# psRefNo - leave ref_no in database
# psRacf - Racf of staff
# return
# ref_no for display, format <office><ref_no><last 3 characters of RACF>
def getDisplayRefNo(psRefNo, psOffice, psRacf):
    return(psOffice + str(psRefNo) + psRacf[-3:])

def  getActualRefNo(psRefNo):
    return(int(psRefNo[3:10]))
    
# get date from string format to mm/dd/yyyy format
# parameter:
# psDateString - Date in String format , i.e. YYYY-MM-DD
# return:
# Date in string format as mm/dd/yyyy
def getMMDDYYYY(psDateString):
    return (datetime.strftime(str2Date(psDateString), "%m/%d/%Y"))

def getWorkdayName(psDateString):
    workdayName = (datetime.strftime(str2Date(psDateString), "%a"))
    return workdayName

# get display leave year
# parameter:
# psYear - leave year in int.
# return:
# leave year period in string, format : "Mar 1, year - Feb 28 (or 29), year"
def getDisplayLeaveYear(psYear):
    if calendar.isleap(psYear):
        return (df['gcYearStartDate'][0] + str(psYear) + " - " + df['gcYearEndDateLeap'][0] + str(psYear + 1))
    else:
        return(df['gcYearStartDate'][0] + str(psYear) + " - " + df['gcYearEndDate'][0] + str(psYear + 1))      

def getLeave(psYear, psLeaveType, psLeaveStatus, psRecord):
    return (list(filter(lambda r: (r["type"].upper() == psLeaveType.upper() and r["applicationStatus"].upper() == psLeaveStatus.upper() and r["year"] == psYear), psRecord["leave_record"])))


def countLeave (psYear, psLeaveType, psLeaveStatus, psRecord):
    leaveDays = 0
    for record in getLeave(psYear, psLeaveType, psLeaveStatus, psRecord):
        for leaveDetails in record["details"]:
            leaveDays += leaveDetails["no_of_workday"]

    return leaveDays

def getYearEntitlement(Year, StaffRecord, LeaveType):
    for rec in StaffRecord['entitlement']:
        if int(rec['year']) == int(Year):
            if LeaveType == 'LVE01':
                return int(rec['annual_entitlement'])
            if LeaveType == 'LVE02':
                return int(rec['casual_entitlement'])

def getYearCarryForward(Year, StaffRecord):
    for rec in StaffRecord['entitlement']:
        if int(rec['year']) == int(Year):
            return int(rec['carry_forward'])

# get leave year period
# parameter:
# psYear - eleave Year
# return:
# list with leave year start date and leave year end date

# Status_code 200: passed
# Status_code 801: Fail to get leave year period


def getLeaveYrPeriod(psYear):
    try:
        yrIndex = json.loads(os.getenv('YEARS')).get('year').index(psYear)
        eleavePeriod = json.loads(os.getenv('YEARS')).get('period')[yrIndex]
        leaveYrStart= datetime.strptime(eleavePeriod.split("-")[0].strip(), "%b %d, %Y")
        leaveYrEnd = datetime.strptime(eleavePeriod.split("-")[1].strip(), "%b %d, %Y")
        leaveYrPeriod = {
            'leaveYrStart': leaveYrStart,
            'leaveYrEnd': leaveYrEnd
        }
        leaveYrPeriodLst = []
        leaveYrPeriodLst.append(leaveYrPeriod)
        return ({"pass": True, "error_message": "", "result":leaveYrPeriodLst, "Status_code": 200})
    except:
        return ({"pass": False, "error_message": "Fail to get leave year period", "result":[], "Status_code": 801})

def chkPeriod(psStartDate, psEndDate, psYear):
    result = getLeaveYrPeriod(psYear)
    if result.get('pass'):
        sDate =  str2Date(psStartDate) 
        eDate = str2Date(psEndDate)
        yrStart = result.get('result')[0].get('leaveYrStart')
        yrEnd = result.get('result')[0].get('leaveYrEnd')
        if sDate >= yrStart and sDate <= yrEnd and eDate >= yrStart and eDate <= yrEnd:
            return ({'pass': True, 'error_message': "", 'result': None})
        else:
            return({'pass': False, 'error_message': "Leave applying is not within the leave year", 'result': None, 'Status_code': 507})
        
    else:
        return ({'pass': False, 'error_message': result.get('error_message'), 'result': None, 'Status_code': 508})

# get leave history of a staff for a particular year
# paramters:
# psRacf - Racf of the user
# psYearStart - Starting Leave Year required
# psYearEnd - Ending Leave Year required
# return:
# list of leave history in the format :
# [{"ref_no": int, "year": int, "type": string, "startDate": string, "startTime": string, "endDate": string, "endTime": string, "status": string, "workDay": int, calendarDay: int, "ldate": datetime, "ltime": string]
def getLeaveHistory(psYearStart, psYearEnd, psRecord):
    leaveHistoryAllLst = [ ]
    yr = psYearStart
    while yr <= psYearEnd:
        for r in psRecord["leave_record"]:
            if r["year"] == yr:
                for d in r["details"]:
                    for p in d["period"]:
                        currRecord = {
                            "ref_no": r["ref_no"],
                            "office": psRecord["staff"]["hr_office"],
                            "racf": psRecord["staff"]["racf"],
                            "staffname": psRecord["staff"]["name"],
                            "empID": psRecord["staff"]["empID"],
                            "dept": psRecord["staff"]["dept"],
                            "position": psRecord["staff"]["position"],
                            "year" : r["year"],
                            "type": r["type"],
                            "sharePointId": r["sharePointId"],
                            "startDate": d["start_date"],
                            "startTime": d["start_time"],
                            "endDate": d["end_date"],
                            "endTime": d["end_time"],
                            "applicationStatus": r["applicationStatus"],
                            "approvalStatus": r["approvalStatus"],
                            "workDay": d["no_of_workday"],
                            "calendarDay": d["no_of_calendarday"],
                            "submitDate": r["submit_date"],
                            "ldate": str2Date(p["ldate"]),
                            "ltime": p["ltime"],
                            "approver1": psRecord["staff"]["approver1"],
                            "approver2": psRecord["staff"]["approver2"],
                            "approver3": psRecord["staff"]["approver3"],
                            "lastUpdate": r["lastUpdate"],
                            "updateDate": r["updateDate"]                            
                        }
                      
                        leaveHistoryAllLst.append (currRecord)
        yr += 1
    return (leaveHistoryAllLst)



# get leave entitlement of a staff for a particular leave type in a particular year
# parameters:
# psRecord : staff record
# psYear : leave year
# psLeaveType : leave type 
# return:
# return leave entitlement, format ["{leaveEntitle": int, "carryForward": int, "forfeoitDate": datetime} ]
def getLeaveEntitlement(psYear, psLeaveTypeAttr, psRecord):
    entitlementLst = [ ]
    entitlement = {
        "leaveEntitle": 0,
        "carryForward": 0,
        "forfeitDate": str2Date("2000-01-01")
    }
    for e in psRecord["entitlement"]:
        if e["year"] == psYear:
            if psLeaveTypeAttr.get("entitlement_field", "") != "":
                leaveEntitle = e.get(psLeaveTypeAttr.get("entitlement_field"), 0)
            else:
                leaveEntitle = 0
            if psLeaveTypeAttr.get("carry_forward_field", "") != "":
                carryForward = e.get(psLeaveTypeAttr.get("carry_forward_field"), 0)
            else:
                carryForward = 0
            if psLeaveTypeAttr.get("forfeit_date_field", "") != "":
                forfeitDate = str2Date(e.get(psLeaveTypeAttr.get("forfeit_date_field"), "2000-01-01"))
            else:
                forfeitDate = str2Date("2000-01-01")
            entitlement = {
                    "leaveEntitle": leaveEntitle,
                    "carryForward": carryForward,
                    "forfeitDate": forfeitDate
            }
    entitlementLst.append(entitlement)  

    return (entitlementLst)

# get all date slot for weekend + holidays within the year period
# parameters:
# psYearStart: Beginning year of the weekend and holidays required
# psYearEnd : Ending year of the weekend and holidays required
# psOffice : Office of the holidays required
# return:
# list of date slot for weekends + holidays within the year period, format : [{"ldate": datetime, "ltime": "AM"/ "PM"}]
def getWeekendHolidays(psYearStart, psYearEnd, psOffice):
    yr = psYearStart
    weekendSlotLst = [ ]
    holidaySlotLst = [ ]
    while yr <= psYearEnd:
        weekendSlotLst = combineTime(weekendSlotLst, getAllWeekend(yr))
        holidaySlotLst = combineTime(holidaySlotLst, getHolidays(yr, psOffice))
        yr += 1
    return (combineTime(weekendSlotLst, holidaySlotLst))

# get no. of working day in the time slot
# parameter:
# psPeriod : time slot list containing leave period in working days
# return:
# no. of work days 
def getWorkDay(psPeriod):
    return (len(psPeriod) / 2)

# expand leave application days to half day timeslot
# check if the leave slot fall into holidays, skip that leave slot if it is 
# check if the leave slot overlapped with those applied before, retrun error if it is
# return whole leave slot otherwise
#parameter : 
#psStartDate : Start Date of leave applying, in string (yyyy-mm-dd) format
#psStartTime : Start Time of leave applying, in string, either "AM" or "PM"
#psEndDate : End Date of leave applying in string (yyyy-mm-dd) format
#psEndTime : End Time of leave applyingm in string, either "AM" or "PM"
#psHolidayLst : List of time slot with holidays and weekends
#psLeaveHistoryLst : List of time slot with leave history in the format [{"ref_no" int, "year" int, "type" string, "status" string, "ldate": datetime, "ltime": "AM" / "PM"}]
#return:
#if no overlap, return list of leave slot for the leave applying in the format :[{"ldate": datetime, "ltime": "AM" / "PM"}]
#if overlap, return empty list
def checkOverlap(psStartDate, psStartTime, psEndDate, psEndTime, psYear, psOffice, psRecord, psApplyingSlotLst, psLeaveType):
    currDate = str2Date(psStartDate)
    currTime = psStartTime
    leaveDtl = [ ]
    weekendHolidaysLst = getWeekendHolidays((psYear - 1), (psYear + 1), psOffice)
    leaveHistoryLst = getLeaveHistory((psYear - 1), (psYear + 1), psRecord)
    #exclude leave that is canceled or rejected
    leaveHistoryLst = list(filter(lambda r: (r["applicationStatus"].upper() != df['gcStatusCancel'][0] and r["applicationStatus"].upper() != df['gcStatusReject'][0]), leaveHistoryLst))
    # loop through the holidays range applied

    while currDate <= str2Date(psEndDate):
        found = [ ]
        # check if the leave slot overlapped with the leave already applied
        overlap = filter(lambda o: (o["ldate"] == currDate) and (o["ltime"] == currTime), leaveHistoryLst)
        if (len(list(overlap))) > 0:
            leaveDtl = [ ]
            leaveDtl = "Leave applied are overlapping each other."
            return leaveDtl
        # check if the leave slot overlapped with the leave applying in different rows
        overlap = filter(lambda o: (o["ldate"] == currDate) and (o["ltime"] == currTime), psApplyingSlotLst)
        if (len(list(overlap))) > 0:
            leaveDtl = [ ]
            leaveDtl = "Leave applying are overlapping each other."
            return leaveDtl        
        # check if the leave slot is in holiday and weekend, if it is not in the holiday list, "found" will be empty and proceed to record in leave detail
        # if the leave slot is in holiday, "found" will not empty and will skip to record in leave detail
        found = list(filter(lambda d: (d["ldate"] == currDate) and (d["ltime"] == currTime), weekendHolidaysLst))
        if len(found) == 1 and (len(leaveDtl)) == 0:
            leaveDtl = [ ]
            leaveDtl = "Leave applying start in Weekends / Holidays"
            return leaveDtl
        if len(found) == 0:
            isHoliday = False
            leaveSlot = { "ldate": currDate, "ltime": currTime, "type": psLeaveType}
            leaveDtl.append(dict(leaveSlot))
        else:
            isHoliday = True
        if currTime.upper() == "PM":
            currDate = currDate + timedelta(1)
            currTime = "AM"
        elif (currDate == str2Date(psEndDate)) and (psEndTime == "AM"):
            break
        else:
            currTime = "PM" 
    if (isHoliday):
        leaveDtl = [ ]
        leaveDtl = "Leave applying end in Weekends / Holidays"
        return leaveDtl   
    return leaveDtl

#check leave balance
#parameters:
#psLeaveEntitle : Leave entitlement for the leave type.
#psCarryForward : Annual leave carry foward from last year
#psForfeitDate:  Forfeit Date for carry forward
#psLeaveHistoryLst : List of time slot with leave history in the format [{"ref_no" int, "year" int, "type" string, "status" string, "ldate": datetime, "ltime": "AM" / "PM"}]
#psLeaveSlotLst : Leave applying
#psLeaveType : Leave type applying
#psYear : Leave year
#return:
#leave balanace after taking counting the leave applying.
def checkBalance(psYear, psLeaveTypeAttr, psRecord, psApplyingLeaveSlotLst):
    leaveEntitleLst = getLeaveEntitlement(psYear, psLeaveTypeAttr, psRecord)
    leaveHistoryLst = getLeaveHistory(psYear, psYear, psRecord)
    leaveHistoryLst = list(filter(lambda r: (r["applicationStatus"].upper() != df['gcStatusCancel'][0] and r["applicationStatus"].upper() != df['gcStatusReject'][0]), leaveHistoryLst))
    beforeForfeit = 0
    afterForfeit = 0
    
    leaveHistoryTypeLst = list(filter(lambda r: (r["type"].upper() == psLeaveTypeAttr.get("leave_type_id").upper() and r["year"] == psYear), leaveHistoryLst))
    # count no. of leave taken before and after the Carry Forward Forfeit Date
    for lve in leaveHistoryTypeLst:
        if lve["ldate"] <= leaveEntitleLst[0]["forfeitDate"]:
            beforeForfeit += 0.5
        else:
            afterForfeit += 0.5
    
    # count no. of leave already taken + applying before and after the Carry Forward Forfeit Date 
    for apply in psApplyingLeaveSlotLst:
        if apply["ldate"] <= leaveEntitleLst[0]["forfeitDate"]:
            beforeForfeit += 0.5
        else:
            afterForfeit += 0.5

    # if total leave taken + applying before the Carry Forward Forfeit Date > total leave entitlement + leave carry forward, return leave balance
    if leaveEntitleLst[0]["leaveEntitle"] + leaveEntitleLst[0]["carryForward"] < beforeForfeit:
        return (leaveEntitleLst[0]["leaveEntitle"] + leaveEntitleLst[0]["carryForward"] - beforeForfeit)

    # if leave remaining after the Forefeit Date >= leave entitlement, set leave entitlement after the Forfeit Date = leave entitlement
    if leaveEntitleLst[0]["leaveEntitle"] + leaveEntitleLst[0]["carryForward"] - beforeForfeit >= leaveEntitleLst[0]["leaveEntitle"]:
        entitleAfterForfeit = leaveEntitleLst[0]["leaveEntitle"]

    # if leave remaining after the Forefeit Date < leave entitlement, set leave entitlement after the Forfeit Date = leave remaining of that year
    else:
        entitleAfterForfeit = leaveEntitleLst[0]["leaveEntitle"] + leaveEntitleLst[0]["carryForward"] - beforeForfeit
    
    # if leave entitlement after the Forfeit Date < leave taken + applying after the Foefeit Date, returm leave balance
    if entitleAfterForfeit < afterForfeit:
        return (entitleAfterForfeit - afterForfeit)

    # if leave taken + applying not exceeding the leave entitlement, return leave balance
    return (entitleAfterForfeit - afterForfeit)

#list all dates for a Week of Day in a specific year.
#parameters :
#psYear : year 
#psDay : Day of week needed, 1 - Monday, 2 - Tuesday, 3 - Wednesday, .... 7- Sunday
#return:
#dates of the day of week required of that year in datetime format.
def alldays(psYear, psDay):
    #d = date(psYear, 1, 1)
    d = str2Date(str(psYear) + "-01-01")
    d += timedelta(days = (psDay - d.isoweekday()) % 7)
    while d.year == psYear:
        yield d
        d += timedelta(days = 7)    

#Get all weekend from year submitted, previous year and next year
#parameter : 
#psYear : year required
#return:
#all sat. and sunday of a required year in the format : [{"ldate": datetime, "ltime": "AM" / "PM"}]
def getAllWeekend(psYear):
    weekendLst = [ ] 
    yr = psYear
    while yr <= psYear:
        dow = 6
        while dow <= 7:
            for d in alldays(yr, dow):
                weekend = {
                    "ldate": d,
                    "ltime": "AM",
                    "type" : "weekend"

                }
                weekendLst.append(weekend)
                weekend = {
                    "ldate": d,
                    "ltime": "PM",
                    "type" : "weekend"
                }
                weekendLst.append(weekend)

            dow += 1
        yr += 1
    return weekendLst

#get all holidays after the year input.  Exclude weekend.
#parameters :
#psYear : Year
#psOffice : Office for the holidays required. 
#return : 
#holidays list of the required year of that office in the format [{"ldate": datetime, "ltime": "AM" / "PM"}]
def getHolidays(psYear, psOffice):
    # convert Date in holiday from string to Date format and exclude weekend.
    holidayLst = list(holidays.find ({ "$and" : [
                                    { "Year":  { "$eq" : psYear } },
                                    { "Office": { "$eq" : psOffice} }
                                ] }
    ) )
    holidaySlotLst = [ ]
    for h in holidayLst:
        if (str2Date(h["Date"])).isoweekday() != 6 and (str2Date(h["Date"])).isoweekday() !=7:
            slot = {
                "ldate": str2Date(h["Date"]),
                "ltime": h["Time"],
                "type": "holiday"
            }
            holidaySlotLst.append (slot)
    return holidaySlotLst


# Combine 2 time slot list and then sort by date and time.
# parameters:
# psLst1 : in the format : [{"ldate": datetime, "ltime": "AM" / "PM"}]
# psLst2 : in the format : [{"ldate": datetime, "ltime": "AM" / "PM"}]
# return : combined list in the format [{"ldate": datetime, "ltime": "AM" / "PM"}]
def combineTime(psLst1, psLst2):
    combinedLst = []

    for s in psLst1:
        slot = {
            "ldate" : s["ldate"],
            "ltime" : s["ltime"],
            "type" : s["type"]
        }
        combinedLst.append(slot)
    
    for s in psLst2:
        slot = {
            "ldate" : s["ldate"],
            "ltime" : s["ltime"],
            "type": s["type"]
            }
        combinedLst.append(slot)

    combinedLst = sorted(combinedLst, key=lambda d: (d['ldate'], d["ltime"]))
    return (combinedLst)

def checkConsecutiveSickLeave (psCombinedSickLeave, psMaxSlNoCert, psApplyingSlotLst):
    slNoCertConsecutiveSlot = 0

    currDate = psCombinedSickLeave[0]["ldate"]
    currTime = psCombinedSickLeave[0]["ltime"]

    currConsecutiveDay = False
    
    for t in psCombinedSickLeave:
        if currDate == t["ldate"]:

            if t["type"] == "LVE05":
                slNoCertConsecutiveSlot += 1
                if t["ldate"] ==  psApplyingSlotLst[0]["ldate"] and currTime == psApplyingSlotLst[0]["ltime"]:
                    currConsecutiveDay = True
                
            currTime = "PM"
        elif (currDate == t["ldate"] + timedelta(-1)) and (currTime == "PM") and (t["ltime"] == "AM"):
            if t["type"] == "LVE05":
                slNoCertConsecutiveSlot += 1
                if t["ldate"] ==  psApplyingSlotLst[0]["ldate"] and currTime == psApplyingSlotLst[0]["ltime"]:
                    currConsecutiveDay = True
            currDate = t["ldate"]
            currTime = t["ltime"]
        else:
            currConsecutiveDay = False
            if t["type"] == "LVE05":
                slNoCertConsecutiveSlot = 1
            else:
                slNoCertConsecutiveSlot = 0
            currDate = t["ldate"]
            currTime = t["ltime"]

        if slNoCertConsecutiveSlot > (psMaxSlNoCert * 2) and currConsecutiveDay:
            return ({"pass": False, "error_message" : "Sick Leave with No Medical Cert. applied is over " + str(psMaxSlNoCert) + " day", "result": None, "daycount": slNoCertConsecutiveSlot - 1, "Status_code": 507})

    return({"pass": True, "error_message": "", "result": None, "daycount": 1, "Status_code": 200}) 

# parameters - leaveHistoryLst, Type
# leaveHistoryLst Example: {'ref_no': 2022001, 'office': 'REG', 'racf': 'NF1BHC', 'staffname': 'BILLY CHAN', 'empID': '00013', 'dept': 'PBT', 'position': 'Regional Analyst Programmer', 'year': 2022, 'type': 'LVE05', 'sharePointId': '', 'startDate': '2022-12-05', 'startTime': 'Full Day', 'endDate': '2022-12-05', 'endTime': 'Full Day', 'applicationStatus': 'PENDING' ...
# Type Example (String): LVE01, LVE02 
def countConsecutiveDaysByType(leaveHistoryLst, ApplyLeaveLst, Type):
    consecutiveSlot = 0

    LeaveHistoryLst = list(filter(lambda r: (r["applicationStatus"].upper() != df['gcStatusCancel'][0] and r["applicationStatus"].upper() != df['gcStatusReject'][0] and r["type"] in Type), leaveHistoryLst))
    LeaveHistoryLst.sort(key=lambda x: x.get('ldate'))
    combinedTimeSlot = combineTime(LeaveHistoryLst, ApplyLeaveLst)

    currDate = combinedTimeSlot[0]["ldate"]
    currTime = combinedTimeSlot[0]["ltime"]

    for t in combinedTimeSlot:
        if currDate == t["ldate"]:
            consecutiveSlot += 1
            currTime = "PM"
        elif (currDate == t["ldate"] + timedelta(-1)) and (currTime == "PM") and (t["ltime"] == "AM"):
            consecutiveSlot += 1
            currDate = t["ldate"]
            currTime = t["ltime"]
        else:
            consecutiveSlot = 1
            currDate = t["ldate"]
            currTime = t["ltime"]

    consecutiveSlot = consecutiveSlot / 2
    
    workDay = getWorkDay(ApplyLeaveLst)
    if workDay > consecutiveSlot:
        consecutiveSlot = workDay
    
    return consecutiveSlot


# check consecutive days of the leave applying and see if it exceeds the consecutive days allowed.
# parameters: 
# psCombinedTimeSlot : list of applying leave, leave already applied, holidays and weekend in the format [{"ldate": datetime, "ltime": "AM" / "PM"}]
# psLimit : max. consecutive days allowed 
# return :
# total consecutive days or the max. consecutive days allowed + 0.5 (when the consective days exceeds the days allowed, it will stop checking and return.)
def checkConsecutiveDays (psYear, psOffice, psRecord, psApplyingSlotLst, psLeaveTypeAttr):
    consecutiveSlot = 0
    groupAttrLst = list(filter(lambda r: (r["groupID"] == psLeaveTypeAttr.get("leave_group")), leaveGroupLst))[0]
    if groupAttrLst.get("max_consecutive_days", "")  != "":
        relatedLveLst = []
        for lve in leaveTypeLst:
            #if lve["leave_group"] == groupAttrLst.get("groupID"):
            if lve["consecutive_days_group"] == psLeaveTypeAttr.get("consecutive_days_group"):
                relatedLveLst.append(lve["leave_type_id"])
    
        leaveHistoryLst = getLeaveHistory((psYear - 1), (psYear + 1), psRecord)
        # if leave is sick leave with cert or sick leave with no cert, check the Consecutive Sick Leave with no cert days
        if groupAttrLst.get("sick_leave", False):
            relatedSlLst = []
            for sl in leaveTypeLst:
                slGrpAttrLst = list(filter(lambda r: (r["groupID"] == sl["leave_group"]), leaveGroupLst))[0]
                if slGrpAttrLst.get("sick_leave", False):
                    relatedSlLst.append(sl["leave_type_id"])
                    if slGrpAttrLst.get("max_consecutive_days", "") != "":
                        maxConsecutiveSlNoCert = slGrpAttrLst.get("max_consecutive_days")
            slLeaveHistoryLst = list(filter(lambda r: (r["applicationStatus"].upper() != df['gcStatusCancel'][0] and r["applicationStatus"].upper() != df['gcStatusReject'][0] and r["type"] in relatedSlLst), leaveHistoryLst))
            combinedTimeSlot = combineTime(slLeaveHistoryLst, psApplyingSlotLst)
            result = checkConsecutiveSickLeave(combinedTimeSlot, maxConsecutiveSlNoCert, psApplyingSlotLst)
            slcount = result.get("daycount")
            if not result.get("pass"):
                return (result)
        leaveHistoryLst = list(filter(lambda r: (r["applicationStatus"].upper() != df['gcStatusCancel'][0] and r["applicationStatus"].upper() != df['gcStatusReject'][0] and r["type"] in relatedLveLst), leaveHistoryLst))
        combinedTimeSlot = combineTime(leaveHistoryLst, psApplyingSlotLst)
        if groupAttrLst.get("consecutive_include_holidays", False):
            weekendHolidaysLst = getWeekendHolidays((psYear - 1), (psYear + 1), psOffice)
            combinedTimeSlot = combineTime (combinedTimeSlot, weekendHolidaysLst)
    
        currDate = combinedTimeSlot[0]["ldate"]
        currTime = combinedTimeSlot[0]["ltime"]

        currConsecutiveDay = False
    
        for t in combinedTimeSlot:
            if currDate == t["ldate"]:
                consecutiveSlot += 1
                if t["ldate"] ==  psApplyingSlotLst[0]["ldate"] and currTime == psApplyingSlotLst[0]["ltime"]:
                    currConsecutiveDay = True
                currTime = "PM"
            elif (currDate == t["ldate"] + timedelta(-1)) and (currTime == "PM") and (t["ltime"] == "AM"):
                consecutiveSlot += 1
                if t["ldate"] ==  psApplyingSlotLst[0]["ldate"] and currTime == psApplyingSlotLst[0]["ltime"]:
                    currConsecutiveDay = True
                currDate = t["ldate"]
                currTime = t["ltime"]                
            else:
                consecutiveSlot = 1
                currDate = t["ldate"]
                currTime = t["ltime"]
                currConsecutiveDay = False
            
            if consecutiveSlot > (groupAttrLst.get("max_consecutive_days", 0) * 2) and currConsecutiveDay:
                return ({"pass": False, "error_message" : "Leave applied is over " + str(groupAttrLst.get("max_consecutive_days")) + " day(s)", "result": None,  "Status_code": 506})

    else:

        return({"pass": True, "error_message": "", "result": None, "Status_code": 200}) 
     
    return ({"pass": True, "error_message": "", "result": None,  "Status_code": 200})

# count the total calendar date.
# input : list of leave applied - psPeriod, list of leave applied + holidays +weekend - psCombinedSlotLst
# total calendar date = 
#   consecutive calendar days (holidays + weekend) before the leave period
#   consecutive calendar days for the leave period +
#   consecutive calendar days (holidays + weekend) after the leave period
# parameter:
# psPeriod - leave slots that need to check for calendar day in the format [{"ldate" : datetime, "ltime": "AM"/ "PM"} ]
# psCombinedSlotLst - leave slots that need to check + all leave applied before + holidays + weekend in the format  [{"ldate" : datetime, "ltime": "AM"/ "PM"} ]
# return :
# No. of calendarDay in int.
def getCalendarDay(psYear, psOffice, psRecord, psPeriod, psLeaveTypeAttr):
    weekendHolidays = getWeekendHolidays((psYear - 1), (psYear + 1), psOffice)
    leaveHistoryLst = getLeaveHistory((psYear - 1), (psYear + 1), psRecord)
    relatedLveLst = []
    for lve in leaveTypeLst:
        if lve["calendar_days_group"] == psLeaveTypeAttr.get("calendar_days_group"):
            relatedLveLst.append(lve["leave_type_id"])
    #leaveHistoryLst = list(filter(lambda r: (r["applicationStatus"].upper() != df['gcStatusCancel'][0] and r["applicationStatus"].upper() != df['gcStatusReject'][0]), leaveHistoryLst))
    leaveHistoryLst = list(filter(lambda r: (r["applicationStatus"].upper() != df['gcStatusCancel'][0] and r["applicationStatus"].upper() != df['gcStatusReject'][0] and r["type"] in relatedLveLst), leaveHistoryLst))
    combinedSlotLst = combineTime(weekendHolidays, leaveHistoryLst)
    combinedSlotLst = combineTime(combinedSlotLst, psPeriod)
    # count the consecutive calendar days (holidays + weekend) before the leave period
    currPosition = next((index for (index, d) in enumerate(combinedSlotLst) if (d["ldate"] == psPeriod[0]["ldate"]) and (d["ltime"] == psPeriod[0]["ltime"])), None)

    calendarDayBefore = 0
    currTime = psPeriod[0]["ltime"]
    currDate = psPeriod[0]["ldate"]
    while currPosition != 0:
        if currDate == combinedSlotLst[currPosition]["ldate"] and currTime == combinedSlotLst[currPosition]["ltime"]:
            currDate = combinedSlotLst[currPosition]["ldate"]
            currTime = combinedSlotLst[currPosition]["ltime"]
            currPosition -= 1
        elif currDate == combinedSlotLst[currPosition]["ldate"] and currTime == "PM" and combinedSlotLst[currPosition]["ltime"] == "AM":
            calendarDayBefore += 0.5
            currDate = combinedSlotLst[currPosition]["ldate"]
            currTime = combinedSlotLst[currPosition]["ltime"]
            currPosition -= 1
        elif currDate == combinedSlotLst[currPosition]["ldate"] + timedelta(days = 1) and currTime == "AM" and combinedSlotLst[currPosition]["ltime"] == "PM":
            calendarDayBefore += 0.5
            currDate = combinedSlotLst[currPosition]["ldate"]
            currTime = combinedSlotLst[currPosition]["ltime"]
            currPosition -= 1
        else:
            break

    # count the consecutive calendar days (holidays + weekend) after the leave period
    currTime = psPeriod[-1]["ltime"]
    currDate = psPeriod[-1]["ldate"]
    currPosition = next((index for (index, d) in enumerate(combinedSlotLst) if (d["ldate"] == psPeriod[-1]["ldate"]) and (d["ltime"] == psPeriod[-1]["ltime"])), None)
    calendarDayAfter = 0
    while currPosition != (len(combinedSlotLst) - 1):
        if currDate == combinedSlotLst[currPosition]["ldate"] and currTime == combinedSlotLst[currPosition]["ltime"]:
            currDate = combinedSlotLst[currPosition]["ldate"]
            currTime = combinedSlotLst[currPosition]["ltime"]
            currPosition += 1      
        elif currDate == combinedSlotLst[currPosition]["ldate"] and currTime == "AM" and combinedSlotLst[currPosition]["ltime"] == "PM":
            calendarDayAfter += 0.5
            currDate = combinedSlotLst[currPosition]["ldate"]
            currTime = combinedSlotLst[currPosition]["ltime"]
            currPosition += 1
        elif currDate == combinedSlotLst[currPosition]["ldate"] - timedelta(days = 1) and currTime == "PM" and combinedSlotLst[currPosition]["ltime"] == "AM":
            calendarDayAfter += 0.5
            currDate = combinedSlotLst[currPosition]["ldate"]
            currTime = combinedSlotLst[currPosition]["ltime"]
            currPosition += 1
        else:
            break
        
    # count the consecutive calendar days for the leave period
    firstPosition = next((index for (index, d) in enumerate(combinedSlotLst) if (d["ldate"] == psPeriod[0]["ldate"]) and (d["ltime"] == psPeriod[0]["ltime"])), None)
    lastPosition = next((index for (index, d) in enumerate(combinedSlotLst) if (d["ldate"] == psPeriod[-1]["ldate"]) and (d["ltime"] == psPeriod[-1]["ltime"])), None)

    calendarDayLeave = ((lastPosition - firstPosition) / 2) + 0.5
    calendarDayTotal = calendarDayBefore + calendarDayLeave + calendarDayAfter


    return (calendarDayTotal)

# get new leave application ref no.
# parameters :
# psLeaveHistoryLst : List of time slot with leave history in the format [{"ref_no" int, "year" int, "type" string, "status" string, "ldate": datetime, "ltime": "AM" / "PM"}]
# psLeaveYear: year
# return:
# if this is the first leave application in the year, return as "leave year" + "001"
# else, return the max ref_no + 1
def getNewRefNo(psYear, psRecord):
    currYearRecordLst = list(filter(lambda r: (r["year"] == psYear), psRecord["leave_record"]))
    
    if len(currYearRecordLst) == 0:
        return (int(str(psYear) + "001"))
    else:
        maxRefNo = max(r["ref_no"] for r in currYearRecordLst)
        return (maxRefNo + 1)


# update / save leave record to database
# parameters:
# psRecord : all records in the collection eleave_dtl in MongoDB
# psID : Object ID of the applicant
# psField : Field that we record to make changes
# psValue : New Value
# return:
# True if update successful
# False if update fail
def updateDB(psRecord, psID, psField, psValue):
    query = { "_id" : psID}
    entry = psRecord.find_one(query)
    entry[psField] = psValue
    try:
        psRecord.update_one(query, { '$set': {psField : entry[psField]}})
        return True
    except:
        return False

def updateDB2(psID, psUpdateLst):
    session = client.start_session(causal_consistency=True)
    session.start_transaction ()
    try: 
        for item in psUpdateLst:
            field = item.get("field")
            value = item.get("value")
            e = eleaveDtl.update_one (
                {"_id": psID },
                { "$set" : { field : value } },
                session=session
            )            
    except Exception as e:
        session.abort_transaction()
        return ({"pass": False, "error_message" : str(e), "result": [ ], "Status_code": 200})
    else:
        session.commit_transaction()
        return ({"pass": True, "error_message": "", "result" : [], "Status_code": 200})
    finally:
        session.end_session()

def checkSSL(host,port,timeout=1):
    sock = socket.socket(socket.AF_INET,socket.SOCK_STREAM) #presumably 
    sock.settimeout(timeout)
    try:
       sock.connect((host,port))
    except:
       return 404
    else:
       sock.close()
       return 250


def Mailer_to_Go(message, sendTo, sendCC):


    # sender
    try: #Heroku
        sender_user = 'noreply'
        sender_email = "@".join([sender_user, current_app.mailertogo_domain])
        sender_name = 'noreply@mmgoverseas.app'
    except: #Local
        sender_email = 'noreply' + "@" + os.environ["MAILERTOGO_DOMAIN"]

    # recipient
    # By Vincent Cheng temp on 11/23/22
    #recipient_email = "ken.yip@macys.com;vincent.cheng@macys.com"
    #recipient_email = "brown.michael.v@gmail.com"
    #recipient_email = "ken.yip@macysinc.onmicrosoft.com;vincent.cheng@macysinc.onmicrosoft.com"    

    #Get recipient domain name
    try: #Heroku
        recipient_domain = current_app.recipient_domain
        local_recipient_domain = current_app.macys_domain
        recipient_email = sendTo.replace(local_recipient_domain, recipient_domain)
    except: #Local
        recipient_domain = current_app.config['recipient_domain']
        recipient_email = sendTo

    #Get recipient cc domain name
    try: #Heroku
        recipient_domain = current_app.recipient_domain
        local_recipient_domain = current_app.macys_domain
        recipient_cc_email = sendCC.replace(local_recipient_domain, recipient_domain)
    except: #Local
        recipient_domain = current_app.config['recipient_domain']
        recipient_cc_email = sendCC

    # subject
    subject = "Eleave System Notification"

    # text body
    body_plain = message

    # html body
    line_break = '\n' #used to replace line breaks with html breaks
    body_html = f'''<html>
        <head></head>
        <body>
        {'<br/>'.join(body_plain.split(line_break))}
        </body>
        </html>'''

    # create message container
    message = MIMEMultipart('alternative')
    message['Subject'] = subject
    message['From'] = sender_email
    message['To'] = recipient_email
    message['Cc'] = recipient_cc_email

    print (message['From'])
    print (message['To'])
    print (message['Cc'])

    # prepare plain and html message parts
    part1 = MIMEText(body_plain, 'plain')
    part2 = MIMEText(body_html, 'html')

    # attach parts to message

    message.attach(part1)
    message.attach(part2)

    # transform recipient to list
    recipient_email = list(recipient_email.split(";"))
    recipient_cc_email = list(recipient_cc_email.split(";"))

    try:
        host = current_app.mailertogo_host
        port = current_app.mailertogo_port
    except:
        host = os.environ["MAILERTOGO_SMTP_HOST"]
        port = os.environ["MAILERTOGO_SMTP_PORT"]
        sc = checkSSL(host, int(port))
        if sc == 404:
            print ("Error Code 404, SMTP connection timeout.")
            quit()
    else:
        pass

    # send the message.
    try:
        server = smtplib.SMTP(host, port)
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(current_app.mailertogo_user, current_app.mailertogo_password)
        server.sendmail(sender_email, (recipient_email + recipient_cc_email), message.as_string())
        server.close()
    except Exception as e:
        server = smtplib.SMTP(host, port)
        server.ehlo()
        server.sendmail(sender_email, (recipient_email + recipient_cc_email), message.as_string())
        server.close()
        print ("Local SMTP Email Sent!")
    else:
        print ("Email sent!")


def sendEmail(psRecord, psRefNo, psApprovalStatus, psAction, psRequest, finalapprover = 1, currentapprover = 1):
    
    leaveContent = list(filter(lambda r: (r["ref_no"] == psRefNo), psRecord["leave_record"]))
    leavePeriod = ""
    for leaveitem in leaveContent[0]["details"]:
        leavePeriod = leavePeriod + leaveitem.get("start_date") + " " + leaveitem.get("start_time") + " " + leaveitem.get("end_date") + " " + leaveitem.get("end_time") + "\n"

    
    # Make email list for sending out to specific recipient by defined cc_general in MongoDB
    cc_general = str(psRecord["staff"]["cc_general"]).replace(",", ";")
    cc_general = cc_general.split(";")
    for index, recipient in enumerate(cc_general):
        cc_general[index] = getStaffRecord(recipient.strip())['staff']["email"]
    cc_general_list = ';'.join(map(str, cc_general))

    sickleave_count = 0

    for leaveitem in leaveContent[0]["details"]:
        try:
            if leaveitem.get("no_of_consective") > sickleave_count: sickleave_count = leaveitem.get("no_of_consective")
        except:
            sickleave_count = 0

    if (int(sickleave_count) >= 2):
        # Make email list for sending out to specific recipient by defined cc_sl2days in MongoDB
        cc_sl2days = str(psRecord["staff"]["cc_sl2days"]).replace(",", ";")
        cc_sl2days = cc_sl2days.split(";")
        for index, recipient in enumerate(cc_sl2days):
            cc_sl2days[index] = getStaffRecord(recipient.strip())['staff']["email"]
        cc_sl2days_list = ';'.join(map(str, cc_sl2days))
    else:
        cc_sl2days_list = ""

    type = ((list(filter(lambda r: (r["leave_type_id"].upper() == leaveContent[0].get("type")), leaveTypeLst))[0]).get("leave_type")) 

    if psRequest == df['gcActionApply'][0] and psAction == df['gcActionApply'][0]:
        sendTo = psRecord["staff"]["email"]
        sendCC = cc_general_list + ";" + cc_sl2days_list
        applicant = psRecord["staff"]["name"]
        message =  '"' + type + '"' + " Ref no. : " + getDisplayRefNo (leaveContent[0].get("ref_no"), psRecord["staff"]["hr_office"], psRecord["staff"]["racf"]) + "\nYour application has been submitted for approval.\n\n Leave Period:\n" + leavePeriod
        try:
            Mailer_to_Go(message, sendTo, sendCC)
        except:
            pass
    elif psRequest == df['gcActionCancel'][0] and psAction == df['gcActionCancel'][0]:
        sendTo = psRecord["staff"]["email"]
        sendCC = cc_general_list + ";" + cc_sl2days_list
        applicant = psRecord["staff"]["name"]
        message =  '"' + type + '"' + " Ref no. : " + getDisplayRefNo (leaveContent[0].get("ref_no"), psRecord["staff"]["hr_office"], psRecord["staff"]["racf"]) + "\nYour cancellation request has been submitted for approval. \n\n LeavePeriod:\n " + leavePeriod
        try:
            Mailer_to_Go(message, sendTo, sendCC)
        except:
            pass               
    if psApprovalStatus == df['gcStatusPending1'][0]:
        sendTo = getStaffRecord(psRecord["staff"]["approver1"])['staff']["email"]
        sendCc = psRecord["staff"]["email"]
        applicant = psRecord["staff"]["name"]
        message =  '"' + type + '"' + " Ref no. : " + getDisplayRefNo (leaveContent[0].get("ref_no"), psRecord["staff"]["hr_office"], psRecord["staff"]["racf"]) + " for " + applicant + " is waiting for your approval. \n\n  Leave Period:\n" + leavePeriod + "\n Link to Approval Center : https://mmgeleave.herokuapp.com/#/ApprovalCenter"
        try:
            Mailer_to_Go(message, sendTo, sendCc)
        except:
            pass
    elif psApprovalStatus == df['gcStatusPending2'][0]:
        sendTo = getStaffRecord(psRecord["staff"]["approver2"])['staff']["email"]
        sendCc = psRecord["staff"]["email"]
        applicant = psRecord["staff"]["name"]
        message =  '"' + type + '"' + " Ref no. : " + getDisplayRefNo (leaveContent[0].get("ref_no"), psRecord["staff"]["hr_office"], psRecord["staff"]["racf"]) +  " for " + applicant + " is waiting for your approval. \n\n  Leave Period:\n" + leavePeriod + "\n Link to Approval Center : https://mmgeleave.herokuapp.com/#/ApprovalCenter"
        try:
            Mailer_to_Go(message, sendTo, sendCc)
        except:
            pass
    elif psApprovalStatus == df['gcStatusPending3'][0]:
        sendTo = getStaffRecord(psRecord["staff"]["approver3"])['staff']["email"]
        sendCc = psRecord["staff"]["email"]
        applicant = psRecord["staff"]["name"]
        message =  '"' + type + '"' + " Ref no. : " + getDisplayRefNo (leaveContent[0].get("ref_no"), psRecord["staff"]["hr_office"], psRecord["staff"]["racf"]) + " for " + applicant + " is waiting for your approval. \n\n  Leave Period:\n" + leavePeriod + "\n Link to Approval Center : https://mmgeleave.herokuapp.com/#/ApprovalCenter"
        try:
            Mailer_to_Go(message, sendTo, sendCc)
        except:
            pass     
    elif psApprovalStatus == df['gcStatusPendingCancel1'][0]:
        sendTo = getStaffRecord(psRecord["staff"]["approver1"])['staff']["email"]
        sendCc = psRecord["staff"]["email"]
        applicant = psRecord["staff"]["name"]
        message =  '"' + type + '"' + " Ref no. : " + getDisplayRefNo (leaveContent[0].get("ref_no"), psRecord["staff"]["hr_office"], psRecord["staff"]["racf"]) + " for " + applicant + " is waiting for your approval. \n\n  Leave Period:\n" + leavePeriod + "\n Link to Approval Center : https://mmgeleave.herokuapp.com/#/ApprovalCenter"
        try:
            Mailer_to_Go(message, sendTo, sendCc)
        except:
            pass
    elif psApprovalStatus == df['gcStatusPending2'][0]:
        sendTo = getStaffRecord(psRecord["staff"]["approver2"])['staff']["email"]
        sendCc = psRecord["staff"]["email"]
        applicant = psRecord["staff"]["name"]
        message =  '"' + type + '"' + " Ref no. : " + getDisplayRefNo (leaveContent[0].get("ref_no"), psRecord["staff"]["hr_office"], psRecord["staff"]["racf"]) + " for " + applicant + " is waiting for your approval. \n\n  Leave Period:\n" + leavePeriod + "\n Link to Approval Center : https://mmgeleave.herokuapp.com/#/ApprovalCenter"
        try:
            Mailer_to_Go(message, sendTo, sendCc)
        except:
            pass
    elif psApprovalStatus == df['gcStatusPending3'][0]:
        sendTo = getStaffRecord(psRecord["staff"]["approver3"])['staff']["email"]
        sendCc = psRecord["staff"]["email"]
        applicant = psRecord["staff"]["name"]
        message =  '"' + type + '"' + " Ref no. : " +getDisplayRefNo (leaveContent[0].get("ref_no"), psRecord["staff"]["hr_office"], psRecord["staff"]["racf"]) + " for " + applicant + " is waiting for your approval. \n\n  Leave Period:\n" + leavePeriod + "\n Link to Approval Center : https://mmgeleave.herokuapp.com/#/ApprovalCenter"
        try:
            Mailer_to_Go(message, sendTo, sendCc)
        except:
            pass
    elif psAction == df['gcActionApprove'][0] and psRequest == df['gcActionApply'][0]:
        sendTo = psRecord["staff"]["email"]
        if currentapprover == finalapprover:
            sendCc = cc_general_list + ";" + cc_sl2days_list
        else:
            sendCc = ""
        applicant = psRecord["staff"]["name"]
        message =   '"' + type + '"' + " Ref no. : " + getDisplayRefNo (leaveContent[0].get("ref_no"), psRecord["staff"]["hr_office"], psRecord["staff"]["racf"]) + "\nYour application has been approved.\n\n Leave Period:\n" + leavePeriod
        try:
            Mailer_to_Go(message, sendTo, sendCc)
        except:
            pass
    elif psAction == df['gcActionReject'][0] and psRequest == df['gcActionApply'][0]:
        sendTo = psRecord["staff"]["email"]
        sendCc = cc_general_list + ";" + cc_sl2days_list
        applicant = psRecord["staff"]["name"]
        message =   '"' + type + '"' + " Ref no. : " + getDisplayRefNo (leaveContent[0].get("ref_no"), psRecord["staff"]["hr_office"], psRecord["staff"]["racf"]) + "\nYour application has been rejected.\n\n Leave Period:\n" + leavePeriod
        try:
            Mailer_to_Go(message, sendTo, sendCc)
        except:
            pass
    elif psAction== df['gcActionApprove'][0] and psRequest == df['gcActionCancel'][0]:
        sendTo = psRecord["staff"]["email"]
        if currentapprover == finalapprover:
            sendCc = cc_general_list + ";" + cc_sl2days_list
        else:
            sendCc = ""
        applicant = psRecord["staff"]["name"]
        message =   '"' + type + '"' + " Ref no. : " + getDisplayRefNo (leaveContent[0].get("ref_no"), psRecord["staff"]["hr_office"], psRecord["staff"]["racf"]) + "\nYour cancellation request has been approved.\n\n Leave Period:\n" + leavePeriod
        try:
            Mailer_to_Go(message, sendTo, sendCc)
        except:
            pass
    elif psAction == df['gcActionReject'][0] and psRequest == df['gcActionCancel'][0]:
        sendTo = psRecord["staff"]["email"]
        sendCc = cc_general_list + ";" + cc_sl2days_list
        applicant = psRecord["staff"]["name"]
        message =   '"' + type + '"' + " Ref no. : " + getDisplayRefNo (leaveContent[0].get("ref_no"), psRecord["staff"]["hr_office"], psRecord["staff"]["racf"]) + "\nYour cancellation request has been rejected.\n\n Leave Period:\n" + leavePeriod
        try:
            Mailer_to_Go(message, sendTo, sendCc)
        except:
            pass
    
# function to apply leave
# parameters:
# psOffice : Office o f that staff for calculating holidays.
# psYear : Annual Leave Year
# psRacf : RACF of the applicant
# psLeaveType : Leave type applying. either :
#           "Annual Leave"
#           "Casual Leave"
#           "Sick Leave - No Medical Cert."
#           "Sick Leave - With Medical Cert."
#           "Work From Home"
# psLeaveLst : Leave period list, coverted into AM, PM.  format as [{"startDate": "2021-07-20", "startTime": "AM", "endDate": "2021-07-20", "endTime": "PM"}]
# psLeaveScreenLst : Leave period list, as at screen showing.  format as [{"startDate": "2021-07-20", "startTime": "Full Day, "endDate": "2021-07-21", "endTime": "AM"}]
# psSuperUser : whether the action is in super use mode.  True / False
# return:
# reject = 0 : leave application, pass = true, no error message.  Leave details will insert into database
# reject = 1 : leave application failed, period overlap found.  pass = false, error message : Leave applied are overlapping each other.
# reject = 2 : leave application failed, consecutive days not pass for Annual Leave and Casual Leave.  pass = false, error message : Leave applied is over 2 weeks.
# reject = 3 : leave application failed, consecutive days not pass for Sick Leave with no cert.  pass = false, error message : Medical certificate is required if sick leave application is more than 1 Day.
# reject = 4 : leave application failed, leave applying > leave balance. pass = false, error message : Not enough days left for the leave.
# reject = 5 : leave application failed, cannot update database.  pass = false, error message : Fail to update database

def applyLeave (psInput):
    getLeaveTypes()
    getLeaveGroups()
    psYear = psInput.get("year", 0)
    psRacf = psInput.get("racf", "")
    psLeaveType = psInput.get("type", "")
    psLeaveLst = psInput.get("applying", "")
    psLeaveScreenLst = psInput.get("applyingScreen", "")
    psUpdateDB = psInput.get("updateDB", True)
    SharePointID = psInput.get("sharePointId")

    # Get Super User 
    try:
        psSuperUser = session["superUser"]
    except:
        psSuperUser = False
    
    if psSuperUser:
        psSuperUser = psInput.get("superUser")
    else:
        psSuperUser = False



    if psYear == 0 or len(psRacf) == 0 or len(psLeaveType) == 0 or len(psLeaveLst) == 0 or len(psLeaveScreenLst) == 0:
        return ({"pass": False, "error_message" : "Incorrect parameters", "result": None, "Status_code": 505})
    staffRecord = getStaffRecord(psRacf)
    leaveHistoryLst = getLeaveHistory(psYear, psYear, staffRecord)
    leaveTypeAttr = (list(filter(lambda r: (r["leave_type_id"].upper() == psLeaveType), leaveTypeLst))[0])
    if not isinstance(staffRecord, dict):
        return ({"pass": False, "error_message" : "Staff Record Not Exist", "result": None, "Status_code": 504}) 
    if len (leaveTypeAttr) == 0:
        return ({"pass": False, "error_message" : "Leave Type Not Found", "result": None, "Status_code": 503}) 
    office = staffRecord["staff"]["office"]
    # applyingSlotLst : list which combine leave application in all rows
    # applyingSlotLstByRow : list which keep leave application row by row.
    applyingSlotLst = [ ]
    applyingSlotLstByRow = [ ]
    overlap = False
    # Loop through each applying leave period
    for rec in psLeaveLst:  
        
        withinYr = chkPeriod (rec["startDate"], rec["endDate"], psYear)
        if not withinYr.get('pass'):
            return({"pass": False, "error_message": withinYr.get('error_message'), "result": None, "Status_code": withinYr.get('Status_code')})
        tmpApplyingSlotLst = checkOverlap(rec["startDate"], rec["startTime"], rec["endDate"], rec["endTime"],  psYear, office, staffRecord, applyingSlotLst, psLeaveType)
        # If no overlap, will get the expanded date slot for leave applying, else leaveSlotLst is empty    
        # put the expanded date slot into applyingSlotLstByRow for saving to DB into separate document.
        #if len(tmpApplyingSlotLst) > 0:
        if isinstance(tmpApplyingSlotLst, list):
            applyingSlotLst = combineTime(applyingSlotLst, tmpApplyingSlotLst)
            applyingSlotLstByRow.append(tmpApplyingSlotLst)
        else:
            overlap = True
            errormsg = tmpApplyingSlotLst
            break    
    if overlap:
        return ({"pass": False, "error_message" : errormsg, "result": None, "Status_code": 502})
    else:
        # Leave type = Annual Leave or Casual Leave :
        # 1. No overlap with
        #  the period already applied
        # 2. Consecutive leave days cannot more than the limits (include annual leave, casual leave, public holidays and weekends), unless leave is applied under superuser mode 
        # 3. Leave applied cannot more than leave entitle + carry forward.
        #if leaveTypeAttr.get("max_consecutive_days",0) > 0:
        result = checkConsecutiveDays(psYear, office, staffRecord, applyingSlotLst, leaveTypeAttr) 
        if not result.get("pass") and not psSuperUser:
            return (result)

        if leaveTypeAttr.get("entitlement_field", "") != "":
            if checkBalance(psYear, leaveTypeAttr, staffRecord, applyingSlotLst) < 0:
                return ({"pass": False, "error_message" : "Not enough days left for the leave.", "result": None, "Status_code": 501})
    
        newRefNo = getNewRefNo(psYear, staffRecord)
        rowNo = 0
        rowDtlLst = [ ]
        for row in applyingSlotLstByRow:
            #noOfCalendarDay = getCalendarDay(psYear, office, staffRecord, row)
            noOfCalendarDay = getCalendarDay(psYear, office, staffRecord, row, leaveTypeAttr)
            noOfWorkDay = getWorkDay (row)
            timeSlotLst = [ ]
            for s in row:
                timeslot = {
                    "ldate" : date2Str(s["ldate"]),
                    "ltime": s["ltime"]
                }
                timeSlotLst.append(timeslot)

            if (row[0]['type']) not in ["LVE04","LVE05"]:
                
                rowDtl = {
                    "start_date": psLeaveScreenLst[rowNo]["startDate"],
                    "start_time": psLeaveScreenLst[rowNo]["startTime"],
                    "end_date": psLeaveScreenLst[rowNo]["endDate"],
                    "end_time": psLeaveScreenLst[rowNo]["endTime"],
                    "no_of_workday": noOfWorkDay,
                    "no_of_calendarday": noOfCalendarDay,             
                    "period" : timeSlotLst
                        }
            elif (row[0]['type']) in ["LVE04","LVE05"]:
                rowDtl = {
                    "start_date": psLeaveScreenLst[rowNo]["startDate"],
                    "start_time": psLeaveScreenLst[rowNo]["startTime"],
                    "end_date": psLeaveScreenLst[rowNo]["endDate"],
                    "end_time": psLeaveScreenLst[rowNo]["endTime"],
                    "no_of_workday": noOfWorkDay,
                    "no_of_calendarday": noOfCalendarDay,             
                    "no_of_consective": countConsecutiveDaysByType(leaveHistoryLst, applyingSlotLst, ["LVE04","LVE05"]),
                    "period" : timeSlotLst
                        }
            rowDtlLst.append(rowDtl)
            rowNo += 1
            approvallist = {
                "approver1": staffRecord['staff']['approver1'],
                "approval_date1": "",
                "approver2": staffRecord['staff']['approver2'],
                "approval_date2": "",                
                "approver3": staffRecord['staff']['approver3'],
                "approval_date3": ""
            }

        newLeaveRecord = {
            "ref_no" : newRefNo,
            "sharePointId" : SharePointID,
            "year" : psYear,
            "type" : psLeaveType,
            "applicationStatus" : df['gcStatusPending'][0],
            "approvalStatus": df['gcStatusPending1'][0],
            "submit_date": date2Str(date.today()),
            "lastUpdate": psRacf,
            "updateDate": date2Str(date.today()),
            "approval": approvallist,
            "details": rowDtlLst
        }
        id = staffRecord["_id"]
        leaveRecord = staffRecord["leave_record"]
        leaveRecord.append(newLeaveRecord)
        updateRecordLst = [ ]
        updateRecord = {
            "field" : "leave_record",
            "value" : leaveRecord,
        }
        updateRecordLst.append(updateRecord)
        if psUpdateDB:
            result = updateDB2(id, updateRecordLst)
        else:
            return ({"pass": True, "error_message" : "VALIDATION MODE.  Data pass validation.  Database NOT updated !", "result": [{"workday": noOfWorkDay, "calendarDay": noOfCalendarDay}], "Status_code": 200})
        if result.get("pass") and psUpdateDB:
            sendEmail(staffRecord, newRefNo, df['gcStatusPending1'][0], df['gcActionApply'][0], df['gcActionApply'][0], 1, 1)

        return (result)                
    
def listLeave (psInput):
    getLeaveTypes()
    psRacf = psInput.get("racf", "")
    psYear = psInput.get("year", 0)

    if len(psRacf) == 0 or psYear == 0:
        return ({"pass": False, "error_message" : "Incorrect parameters", "result": None, "Status_code": 505})
    displayLeaveHistoryDtl = [ ]
    displayLeaveHistoryHdr = [ ]
    
    staffRecord = getStaffRecord(psRacf)
    if not isinstance(staffRecord, dict):
        return ({"pass": False, "error_message" : "Staff Record Not Exist", "result": None, "Status_code": 504}) 

    for lveType in leaveTypeLst:
        leaveTypeHdr = {
            "leaveType": lveType.get("leave_type"),
            "leaveTypeId": lveType.get("leave_type_id"),
            "taken" : countLeave(psYear, lveType.get("leave_type_id"), df['gcStatusApproved'][0], staffRecord),
            "pending": countLeave(psYear, lveType.get("leave_type_id"), df['gcStatusPending'][0], staffRecord),
            "balance": checkBalance(psYear, lveType, staffRecord, [])
        }
        displayLeaveHistoryHdr.append(leaveTypeHdr)    
        
    leaveHistoryLst = getLeaveHistory(psYear, psYear, staffRecord)
    currRefNo = 0
    currStartDate = ""
    currStartTime = ""
    for lve in leaveHistoryLst:
        if (currRefNo == lve["ref_no"] and currStartDate != lve["startDate"] and currStartTime != lve["startTime"]) or (currRefNo != lve["ref_no"]):
            displayLeaveRecord = {
                    "submitDate":  getMMDDYYYY(lve["submitDate"]),
                    "refNo": getDisplayRefNo(lve["ref_no"], lve["office"], lve["racf"]),
                    "office": lve["office"],
                    "staffname": lve["staffname"],
                    "empID": lve["empID"],
                    "dept": lve["dept"],
                    "position": lve["position"],
                    "type_id": lve["type"],
                    "sharePointId": lve["sharePointId"],
                    "type" : list(filter(lambda r: (r["leave_type_id"].upper() == lve["type"]), leaveTypeLst))[0].get("leave_type"),
                    "year": getDisplayLeaveYear(lve["year"]),
                    "leaveFrom": getMMDDYYYY(lve["startDate"]),
                    "startPeriod": lve["startTime"],
                    "leaveTo": getMMDDYYYY(lve["endDate"]),
                    "endPeriod": lve["endTime"],
                    "workday": lve["workDay"],
                    "calendarDay": lve["calendarDay"],
                    "applicationStatus": lve["applicationStatus"],
                    "approver1": lve["approver1"],
                    "approver2": lve["approver2"],
                    "approver3": lve["approver3"],
                    "approvalStatus" : lve["approvalStatus"],
                    "lastUpdate": lve["lastUpdate"],
                    "updateDate": getMMDDYYYY(lve["updateDate"])
            }
            displayLeaveHistoryDtl.append(displayLeaveRecord)
            currRefNo = lve["ref_no"]
            currStartDate = lve["startDate"]
            currStartTime = lve["endDate"]

    fullRecord = {
        "header":  displayLeaveHistoryHdr,
        "details": displayLeaveHistoryDtl
    }
    leaveRecordLst = [ ]
    leaveRecordLst.append (fullRecord) 

    if len(leaveRecordLst) == 0:
        return ({"pass": True, "error_message" : None, "result": [], "Status_code": 200}) 
    else:
        return ({"pass": True, "error_message" : None, "result": leaveRecordLst, "Status_code": 200}) 


def listApprove(psInput):
  

    getLeaveTypes()
    psApprover = psInput.get("racf", "")
    print ('listApprove',  psApprover)

    if len(psApprover) == 0:
        return ({"pass": False, "error_message" : "Incorrect parameters", "result": None, "Status_code": 505})

    approvalRecordLst = [ ]

    i = 1
    while i <= 6:
        if i == 1:
            approver = "staff.approver1"
            pendingStatus = df['gcStatusPending1'][0]
        elif i == 2:
            approver = "staff.approver2"
            pendingStatus = df['gcStatusPending2'][0]
        elif i == 3:
            approver = "staff.approver3"
            pendingStatus = df['gcStatusPending3'][0]
        elif i == 4:
            approver = "staff.approver1"
            pendingStatus = df['gcStatusPendingCancel1'][0]
        elif i == 5:
            approver = "staff.approver2"
            pendingStatus = df['gcStatusPendingCancel2'][0]
        else:
            approver = "staff.approver3"
            pendingStatus = df['gcStatusPendingCancel3'][0]
        #tmpApproverLst = [ ]
        staffRecord = list(eleaveDtl.find ( {approver : { '$regex' : psApprover, '$options' : "i"} , "staff.status": { '$regex': "ACTIVE", '$options': "i"} } ) )
        for rec in staffRecord:
            staff = rec["staff"]["name"]
            racf = rec["staff"]["racf"]
            office = rec["staff"]["hr_office"]
            pendingLst = list(filter(lambda r: (r["approvalStatus"] == pendingStatus), rec["leave_record"]))
            for record in pendingLst:
                leaveDetailsLst = [ ]
                for details in record["details"]:
                    # get rows for each leave application
                    leaveDetails = {
                            "startDate": details["start_date"],
                            "startTime": details["start_time"],
                            "endDate": details["end_date"],
                            "endTime": details["end_time"],
                            "workday": details["no_of_workday"],
                            "calendarDay": details["no_of_calendarday"]
                    }
                    leaveDetailsLst.append(leaveDetails)
                # put a single leave application , by ref_no, into a dict
                leaveRecord = {
                    "staff": staff,
                    "racf": racf,
                    "ref_no": getDisplayRefNo(record["ref_no"], office, racf),
                    "sharePointId": record["sharePointId"],
                    "type_id": record["type"],
                    "type": list(filter(lambda r: (r["leave_type_id"].upper() == record["type"]), leaveTypeLst))[0].get("leave_type"),
                    "approvalStatus": pendingStatus,
                    "details": leaveDetailsLst
                }
                # add that leave application into the whole list (temporary list)
                approvalRecordLst.append(leaveRecord)
        i += 1

        approvalRecordLst = sorted(approvalRecordLst, key=lambda d: (d["approvalStatus"], d["staff"], d["racf"], d["ref_no"]))
    return ({"pass": True, "error_message" : None, "result": approvalRecordLst, "Status_code": 200}) 

def changeStatus(psInput):
    psRefNo = psInput.get("refNo", 0)
    psRacf = psInput.get("racf", "")
    psAction = psInput.get("action","")

    approval_index = 1
    approver = ""
        ## Super User get session data
    try:
        psSuperUser = session["superUser"]
    except:
        psSuperUser = psInput.get("superUser", False)

    if psRefNo == 0 or len(psRacf) == 0 or len(psAction) == 0:
        return ({"pass": False, "error_message" : "Incorrect parameters", "result": None, "Status_code": 505})        
    refNo = getActualRefNo(psRefNo)
    applicantRacf = "NF1" + psRefNo[-3:]
    staffRecord = getStaffRecord(applicantRacf)
    if not isinstance(staffRecord, dict):
        return ({"pass": False, "error_message" : "Staff Record Not Exist", "result": None, "Status_code": 504}) 

    max_approver = 1
    ## Find final approver
    if len(staffRecord["staff"]["approver2"]) != 0: max_approver = 2
    if len(staffRecord["staff"]["approver3"]) != 0: max_approver = 3

    leaveRecord = [(idx, record) for idx, record in enumerate(staffRecord["leave_record"]) if record["ref_no"] == refNo]
    currApplicationStatus = (leaveRecord[0][1]["applicationStatus"])
    currApprovalStatus = (leaveRecord[0][1]["approvalStatus"])

    ## Get work days
    no_of_workdays = (leaveRecord[0][1]["details"][0]["no_of_workday"])

    index = leaveRecord[0][0]
    newApplicationStatus = currApplicationStatus
    if currApplicationStatus == df['gcStatusReject'][0] or currApplicationStatus == df['gcStatusCancel'][0]:
        return ({"pass": False, "error_message" : "Current Status cannot be changed.", "result": None, "Status_code": 603})

    # Action - Cancel
    if psAction == df['gcActionCancel'][0]:
        if psRacf != applicantRacf and not psSuperUser:
            return ({"pass": False, "error_message" : "Only applicant himself / herself can cancel leave.", "result": None, "Status_code": 604})
        if currApprovalStatus == df['gcStatusPendingCancel1'][0] or currApprovalStatus == df['gcStatusPendingCancel2'][0] or currApprovalStatus == df['gcStatusPendingCancel3'][0]:
            return ({"pass": False, "error_message" : "Leave cancel already submitted and waiting for approval.", "result": None, "Status_code": 605})
        firstLeaveDate = str2Date ("9999-12-31")
        for row in leaveRecord[0][1]["details"]:
            if str2Date(row["start_date"]) < firstLeaveDate:
                firstLeaveDate = str2Date(row["start_date"])
        if datetime.today() > firstLeaveDate and not psSuperUser:
            return ({"pass": False, "error_message" : "Cannot cancel leave in the past period.", "result": None, "Status_code": 606})
        newApplicationStatus = currApplicationStatus
        newApprovalStatus = df['gcStatusPendingCancel1'][0]
        updateBy = psRacf
        emailRequest = df['gcActionCancel'][0]

    # Action - Approve     
    elif psAction == df['gcActionApprove'][0]:
        updateBy = psRacf
        if currApprovalStatus == df['gcStatusPending1'][0]:
            approval_index = 1
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver1"] and not psSuperUser:             
               return ({"pass": False, "error_message" : "User is not the first approver.", "result": None, "Status_code": 607}) 
            emailRequest = df['gcActionApply'][0]
            #***
            if len(staffRecord["staff"]["approver2"]) == 0:
                newApplicationStatus = df['gcStatusApproved'][0]
                newApprovalStatus = df['gcStatusApproved'][0]
            else:
                newApprovalStatus = df['gcStatusPending2'][0]
                                
        elif currApprovalStatus == df['gcStatusPending2'][0]:
            approval_index = 2
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver2"] and not psSuperUser:
               return ({"pass": False, "error_message" : "User is not the second approver.", "result": None, "Status_code": 608}) 
            emailRequest = df['gcActionApply'][0]
            if len(staffRecord["staff"]["approver3"]) == 0:
                newApplicationStatus = df['gcStatusApproved'][0]
                newApprovalStatus = df['gcStatusApproved'][0]
            else:
                newApprovalStatus = df['gcStatusPending3'][0]
        elif currApprovalStatus == df['gcStatusPending3'][0]:
            approval_index = 3
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver3"] and not psSuperUser:
               return ({"pass": False, "error_message" : "User is not the third approver.", "result": None, "Status_code": 609}) 
            emailRequest = df['gcActionApply'][0]
            newApplicationStatus = df['gcStatusApproved'][0]
            newApprovalStatus = df['gcStatusApproved'][0]
        elif currApprovalStatus == df['gcStatusPendingCancel1'][0]:
            approval_index = 1
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver1"] and not psSuperUser:
               return ({"pass": False, "error_message" : "User is not the first approver.", "result": None, "Status_code": 607})
            emailRequest = df['gcActionCancel'][0]        
            if len(staffRecord["staff"]["approver2"]) == 0:
                newApplicationStatus = df['gcStatusCancel'][0]
                newApprovalStatus = df['gcStatusCancel'][0]
            else:
                newApprovalStatus = df['gcStatusPendingCancel2'][0]
        elif currApprovalStatus == df['gcStatusPendingCancel2'][0]:
            approval_index = 2
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver2"] and not psSuperUser:
               return ({"pass": False, "error_message" : "User is not the second approver.", "result": None, "Status_code": 608})   
            emailRequest = df['gcActionCancel'][0]
            if len(staffRecord["staff"]["approver3"]) == 0:
                newApplicationStatus = df['gcStatusCancel'][0]
                newApprovalStatus = df['gcStatusCancel'][0]
            else:
                newApprovalStatus = df['gcStatusPendingCancel3'][0]
        elif currApprovalStatus == df['gcStatusPendingCancel3'][0]:
            approval_index = 3
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver3"] and not psSuperUser:
               return ({"pass": False, "error_message" : "User is not the third approver.", "result": None, "Status_code": 609}) 
            emailRequest = df['gcActionCancel'][0]
            newApplicationStatus = df['gcStatusCancel'][0]
            newApprovalStatus = df['gcStatusCancel'][0]
        elif currApprovalStatus == df['gcStatusApproved'][0]:
            return ({"pass": False, "error_message" : "Leave already approved.", "result": None, "Status_code": 610}) 
        else:
            return ({"pass": False, "error_message" : "Incorrect action", "result": None, "Status_code": 611}) 
    #Action = Reject
    else:
        updateBy = psRacf 
        if currApprovalStatus == df['gcStatusPending1'][0]:
            approval_index = 1
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver1"] and not psSuperUser:
                return ({"pass": False, "error_message" : "User is not the first approver.", "result": None, "Status_code": 607})
            emailRequest = df['gcActionApply'][0]
            newApplicationStatus = df['gcStatusReject'][0]
            newApprovalStatus = df['gcStatusReject'][0]
        elif currApprovalStatus == df['gcStatusPending2'][0]:
            approval_index = 2
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver2"] and not psSuperUser:
                return ({"pass": False, "error_message" : "User is not the second approver.", "result": None, "Status_code": 608})
            emailRequest = df['gcActionApply'][0]
            newApplicationStatus = df['gcStatusReject'][0]
            newApprovalStatus = df['gcStatusReject'][0]
        elif currApprovalStatus == df['gcStatusPending3'][0]:
            approval_index = 3
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver3"] and not psSuperUser:
                return ({"pass": False, "error_message" : "User is not the third approver.", "result": None, "Status_code": 609})
            emailRequest = df['gcActionApply'][0]
            newApplicationStatus = df['gcStatusReject'][0]
            newApprovalStatus = df['gcStatusReject'][0]      
        elif currApprovalStatus == df['gcStatusPendingCancel1'][0]:
            approval_index = 1
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver1"] and not psSuperUser:
                return ({"pass": False, "error_message" : "User is not the first approver.", "result": None, "Status_code": 607})
            emailRequest = df['gcActionCancel'][0]
            newApplicationStatus = currApplicationStatus
            newApprovalStatus = currApplicationStatus
        elif currApprovalStatus == df['gcStatusPendingCancel2'][0]:
            approval_index = 2
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver2"] and not psSuperUser:
                return ({"pass": False, "error_message" : "User is not the second approver.", "result": None, "Status_code": 608})
            emailRequest = df['gcActionCancel'][0]
            newApplicationStatus = currApplicationStatus
            newApprovalStatus = currApplicationStatus
        elif currApprovalStatus == df['gcStatusPendingCancel3'][0]:
            approval_index = 3
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver3"] and not psSuperUser:
                return ({"pass": False, "error_message" : "User is not the third approver.", "result": None, "Status_code": 609})
            emailRequest = df['gcActionCancel'][0]
            newApplicationStatus = currApplicationStatus
            newApprovalStatus = currApplicationStatus

    id = staffRecord["_id"]
    updateStatusLst = [ ]
    updateStatus = {
        "field": "leave_record." + str(index) + ".applicationStatus",
        "value": newApplicationStatus
    }
    updateStatusLst.append(updateStatus)     
    updateStatus = {
        "field": "leave_record." + str(index) + ".approvalStatus",
        "value": newApprovalStatus
    }
    updateStatusLst.append(updateStatus)
    updateStatus = {
        "field": "leave_record." + str(index) + ".lastUpdate",
        "value": updateBy
    }
    updateStatusLst.append(updateStatus)
    updateStatus = {
        "field": "leave_record." + str(index) + ".updateDate",
        "value": date2Str(date.today())
    }
    updateStatusLst.append(updateStatus)
    # Update leave record data in MongoDB
    result = updateDB2(id, updateStatusLst)

    updateApproval = {
        "field": "leave_record." + str(index) + ".approval" + ".approver" + str(approval_index),
        "value": approver
    }

    updateApprovalLst = [ ]
    updateApprovalLst.append(updateApproval)

    # Get local time from browser and convert to MongoDB Datetime format YYYY-mm-dd
    date_input = datetime.strptime(psInput['localTime'], '%a %b %d %Y')
    new_date = date_input.strftime('%Y-%m-%d')

    updateApproval = {
        "field": "leave_record." + str(index) + ".approval" + ".approval_date" + str(approval_index),
        "value":  new_date
    }
    updateApprovalLst.append(updateApproval)

    # Update approver record data in MongoDB
    approval_result = updateDB2(id, updateApprovalLst)

    
    if result.get("pass") and approval_result.get("pass"):
        staffRecord = getStaffRecord(applicantRacf)
        sendEmail (staffRecord, refNo, newApprovalStatus, psAction, emailRequest, max_approver, approval_index)
    return (result)  

 
def listApprovedLeaveByYear(psInput):
    
    getLeaveTypes()
    psUser = psInput.get("racf", "")

    # Get Super User 
    try:
        superUser = session["superUser"]
    except:
        superUser = False
    
    if superUser:
        superUser = psInput.get("superUser")
    else:
        superUser = False

    # Get Year
    try:
        years = (json.loads(current_app.config['YEARS'])).get('year')
    except:
        years_str = os.environ['YEARS']      
        years = eval(years_str)
        years = pd.DataFrame(data=years)
        years = years['year'].tolist()

    if len(psUser) == 0:
        return ({"pass": False, "error_message" : "Incorrect parameters", "result": None, "Status_code": 505})

    approvalRecordLst = [ ]


    #tmpApproverLst = [ ]
    staffRecord = list(eleaveDtl.find ( {"staff.racf" : { '$regex' : psUser, '$options' : "i"} , "staff.status": { '$regex': "ACTIVE", '$options': "i"} } ) )
    for rec in staffRecord:
        staff = rec["staff"]["name"]
        racf = rec["staff"]["racf"]
        office = rec["staff"]["hr_office"]
        leaveappliedLst = list(filter(lambda r: (r["approvalStatus"] == df['gcStatusApproved'][0]), rec["leave_record"]))
        for record in leaveappliedLst:
            #print (record)
            leaveDetailsLst = [ ]
            ##if record['year'] == psYear:
            ## Added by Vincent to allow multiple years 
            if record['year'] in years:
                for details in record["details"]:
                    #print (details)
                    # get rows for each leave application
                    leaveDetails = {
                            "startDate": details["start_date"],
                            "startTime": details["start_time"],
                            "endDate": details["end_date"],
                            "endTime": details["end_time"],
                            "workday": details["no_of_workday"],
                            "calendarDay": details["no_of_calendarday"]
                                   }
                    #Only show the the leave date after today date if not super user
                    # Get local time from browser and convert to MongoDB Datetime format YYYY-mm-dd
                    date_input = datetime.strptime(psInput['localTime'], '%a %b %d %Y')

                    if (((datetime.strptime(details["start_date"],'%Y-%m-%d').date() - date_input.date()).days) >= 1 and not superUser) or (superUser):
                        leaveDetailsLst.append(leaveDetails)
                    # put a single leave application , by ref_no, into a dict
                leaveRecord = {
                        "staff": staff,
                        "racf": racf,
                        "ref_no": getDisplayRefNo(record["ref_no"], office, racf),
                        "type_id": record["type"],
                        "type": list(filter(lambda r: (r["leave_type_id"].upper() == record["type"]), leaveTypeLst))[0].get("leave_type"),
                        "approvalStatus": df['gcStatusApproved'][0],
                        "details": leaveDetailsLst
                              }
                    # add that leave application into the whole list (temporary list)
                if (((datetime.strptime(details["start_date"],'%Y-%m-%d').date() - date_input.date()).days) >= 1 and not superUser) or (superUser):
                    approvalRecordLst.append(leaveRecord)
        approvalRecordLst = sorted(approvalRecordLst, key=lambda d: (d["approvalStatus"], d["staff"], d["racf"], d["ref_no"]))
    
    
    return ({"pass": True, "error_message" : None, "result": approvalRecordLst, "Status_code": 200}) 

#@app.route("/api/listleave", methods=['POST'])
#@app.route("/")
def apiListLeave():
    psInput = request.get_json()
    result = listLeave(psInput)
    try: 
        return jsonify(result), result['Status_code'] # APP
    except:
        return jsonify(result) # postman


# Status_code 200: passed
# Status_code 51: Fail to generate Leave Summary

@eleave.route("/api/printsummary", methods=['POST'])
@checkLogged.check_logged
def apiPrintSummary():
 
    para = json.loads(request.headers['parameters'])                        
    psInput =  {'year': para['year'], 'racf': para['racf']}    

    result = listLeave(psInput)
 
    rpt = reportMap.find_one ( { "report": "Leave Summary"} )
    
    if (result.get("pass")): 
        hdr = result.get("result")[0]["header"]
        alTaken = 0
        alPending = 0
        alBalance = 0
        clTaken = 0
        clPending = 0
        clBalance = 0
        slTaken = 0
        slPending = 0
        slBalance = 0
        for lve in leaveTypeLst:
            hdrData =  list(filter(lambda r: (r["leaveTypeId"].upper() == lve["leave_type_id"]), hdr))[0]
            if lve["leave_type_id"] == "LVE01":
                alTaken = hdrData["taken"]
                alPending = hdrData["pending"]
                alBalance = hdrData["balance"]
            elif lve["leave_type_id"] == "LVE02":
                clTaken = hdrData["taken"]
                clPending = hdrData["pending"]
                clBalance = hdrData["balance"]
            elif lve["leave_type_id"] == "LVE04":
                slTaken = hdrData["taken"] + slTaken
                slPending = hdrData["pending"] + slPending
                slBalance = hdrData["balance"] + slBalance
            elif lve["leave_type_id"] == "LVE05":
                slTaken = hdrData["taken"] + slTaken
                slPending = hdrData["pending"] + slPending
                slBalance = hdrData["balance"] + slBalance
        rptDtlLst = [ ]                    
        for record in result.get("result"):
            for dtl in record["details"]:
                rptDtl = {
                    "submitDate": dtl.get("submitDate"),
                    "refNo": dtl.get("refNo"),
                    "office": dtl.get("office"),
                    "staffname": dtl.get("staffname"),
                    "empID": dtl.get("empID"),
                    "dept": dtl.get("dept"),
                    "type": dtl.get("type"),
                    "year": dtl.get("year"),
                    "leaveFrom": dtl.get("leaveFrom"),
                    "startPeriod": dtl.get("startPeriod"),
                    "leaveTo": dtl.get("leaveTo"),
                    "endPeriod": dtl.get("endPeriod"),
                    "workday": dtl.get("workday"),
                    "calendarDay": dtl.get("calendarDay"),
                    "applicationStatus": dtl.get("applicationStatus"),
                    "lastUpdate": dtl.get("lastUpdate"),
                    "updateDate": dtl.get("updateDate")        
                }
                rptDtlLst.append(rptDtl)
            # added by Vincent Cheng on 11/23 
            try:
                if record["details"][0].get("year"):
                    pass
            except:            
                return jsonify({"error_message" : "Sorry, we failed to generate Leave Summary.  Perhaps no data for the year"}), 501    
                
    
            report = {
                "hdrCalendarYear": record["details"][0].get("year"),
                "hdrUser": record["details"][0].get("staffname") + "\nLeave Application Summary",
                "hdrALTaken": alTaken,
                "hdrALPending": alPending,
                "hdrALBalance": alBalance,
                "hdrSLTaken": slTaken,
                "hdrSLPending": slPending,
                "hdrCLTaken": clTaken,
                "hdrCLPending": clPending,
                "hdrCLBalance": clBalance,
                "dtl": rptDtlLst
                
            }
        #filename when using in Heroku:
        fs = gridfs.GridFS(db)
        wb = load_workbook(filename=BytesIO(fs.get(ObjectId(rpt["file"]["fileObj"])).read()))

        # filename in development:
        #wb = load_workbook(filename=rpt["file"]["fileName"])
        ws = wb[rpt["file"]["wsName"]]
        
        result = genReport(ws, report, rpt)
        if result[1] == 0:
            out = BytesIO()
            wb.save(out)
            out.seek(0)        
 
            #wb.save(filename="F:\mmgapp\dev\eleave\output\LeaveSummary.xlsx")
            wb.close()            
            print('sending file...')
            return send_file(out,  attachment_filename='a_file.xls', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')            
        else:            
            return jsonify({"error_message" : "Sorry, we failed to generate Leave Summary.  Perhaps no data for the year"}), 501    
        

@eleave.route("/api/printapply", methods=['POST'])
@checkLogged.check_logged
def apiPrintApply():

    # Get RACF and Ref_no from frontend
    para = json.loads(request.headers['parameters'])
    ref = para['ref']
    racf = para['racf']

    # Get Staff record for output
    StaffRecord = getStaffRecord(racf)

    # Approver 
    if len(StaffRecord['staff']['approver1']) > 1:
        approver1 = getStaffRecord(StaffRecord['staff']['approver1'])['staff']['name']
    else:
        approver1 = ""
    if len(StaffRecord['staff']['approver2']) > 1:
        approver2 = getStaffRecord(StaffRecord['staff']['approver2'])['staff']['name']
    else:
        approver2 = ""
    if len(StaffRecord['staff']['approver3']) > 1:
        approver3 = getStaffRecord(StaffRecord['staff']['approver3'])['staff']['name']
    else:
        approver3 = ""

    # Find leave list by racf and ref
    ref_no = ref.replace(StaffRecord['staff']['hr_office'],"") # remove hr office in reference no
    ref_no = ref_no.replace(StaffRecord['staff']['racf'][-3:],"") # remove staff racf in reference no

    # Get leave balance 
    getLeaveTypes()

    approvalRecordLst = [ ]

    for rec in StaffRecord['leave_record']:

        #Select exact application by reference number
        if rec['ref_no'] == int(ref_no):
            leaveDetailsLst = [ ]
            for details in rec["details"]:
                # get rows for each leave application
                leaveDetails = {
                    "startDate": details["start_date"],
                    "startTime": details["start_time"],
                    "endDate": details["end_date"],
                    "endTime": details["end_time"],
                    "workday": details["no_of_workday"],
                    "calendarDay": details["no_of_calendarday"]
                    }
                leaveDetailsLst.append(leaveDetails)

                # Check the balance from Thomas function
                displayLeaveHistoryHdr = [ ]
                for lveType in leaveTypeLst:
                    leaveTypeHdr = {
                                    "leaveType": lveType.get("leave_type"),
                                    "leaveTypeId": lveType.get("leave_type_id"),
                                    "taken" : countLeave(rec['year'], lveType.get("leave_type_id"), df['gcStatusApproved'][0], StaffRecord),
                                    "pending": countLeave(rec['year'], lveType.get("leave_type_id"), df['gcStatusPending'][0], StaffRecord),
                                    "balance": checkBalance(rec['year'], lveType, StaffRecord, [])
                                    }
                    displayLeaveHistoryHdr.append(leaveTypeHdr)
                
            # Summarize the number of balance
            if rec['type'] == 'LVE01':
                DaysOfApproved = displayLeaveHistoryHdr[0]['taken']
                DaysOfPending = displayLeaveHistoryHdr[0]['pending']
                DaysOfleft = displayLeaveHistoryHdr[0]['balance']
                DaysOfCarryForward = getYearCarryForward(rec['year'], StaffRecord)
                DaysOfEntitlement = str(DaysOfCarryForward) + " (" + str(int(rec['year']-1)) + ") " + "+ " + str(getYearEntitlement(rec['year'], StaffRecord, rec['type'])) + " (" + str(int(rec['year'])) + ") "
            elif rec['type'] == 'LVE02':
                DaysOfApproved = displayLeaveHistoryHdr[1]['taken']
                DaysOfPending = displayLeaveHistoryHdr[1]['pending']
                DaysOfleft = displayLeaveHistoryHdr[1]['balance']
                DaysOfCarryForward = 0
                DaysOfEntitlement = str(getYearEntitlement(rec['year'], StaffRecord, rec['type'])) + " (" + str(int(rec['year'])) + ") "
            else:
                DaysOfApproved = "N/A"
                DaysOfPending = "N/A"
                DaysOfleft = "N/A"
                DaysOfEntitlement = "N/A"

            # Display Office Name
            if StaffRecord['staff']['office'] == "HKG": str_officeHeader = "Hong Kong"
            elif StaffRecord['staff']['office'] == "DEL": str_officeHeader = "India"
            elif StaffRecord['staff']['office'] == "FLR": str_officeHeader = "Italy"
            elif StaffRecord['staff']['office'] == "TPE": str_officeHeader = "Taiwan"

            get_approver1 = ""
            get_pos_approver1 = ""
            get_approver2 = ""
            get_pos_approver2 = ""
            get_approver3 = ""
            get_pos_approver3 = ""

            if len(str(rec['approval']['approver1'])) > 0:
                get_approver1 = getStaffRecord(rec['approval']['approver1'])['staff']['name']
                get_pos_approver1 = getStaffRecord(rec['approval']['approver1'])['staff']['position']

            if len(str(rec['approval']['approver2'])) > 0:
                get_approver2 = getStaffRecord(rec['approval']['approver2'])['staff']['name']
                get_pos_approver2 = getStaffRecord(rec['approval']['approver2'])['staff']['position']

            if len(str(rec['approval']['approver3'])) > 0:
                get_approver3 = getStaffRecord(rec['approval']['approver3'])['staff']['name']
                get_pos_approver3 = getStaffRecord(rec['approval']['approver3'])['staff']['position']

            if rec['sharePointId'] == "":
                DissharePointid = ""
            else:
                DissharePointid = "(" + str(rec['sharePointId']) + ")"

            try:
                TakenApproved = int(DaysOfApproved + DaysOfPending)
            except:
                TakenApproved = "NA"

            # Go back to build the structure for excel output file
            # Array item label must be the same as MongoDB cell field in fileDrectory
            leaveRecord = {
                "staff": StaffRecord['staff']['name'],
                "racf": racf,
                "position": StaffRecord['staff']['position'],
                "dept": StaffRecord['staff']['dept'],
                "date_joined": StaffRecord['staff']['date_join'],
                "ref_no": ref,
                "sharePointid": DissharePointid,
                "approver1": get_approver1,
                "approver_pos1": get_pos_approver1,
                "approval_date1": rec['approval']['approval_date1'],
                "approver2": get_approver2,
                "approver_pos2": get_pos_approver2,
                "approval_date2": rec['approval']['approval_date2'],
                "approver3": get_approver3,
                "approver_pos3": get_pos_approver3,
                "approval_date3": rec['approval']['approval_date3'],
                "NoDaysEntitlement": DaysOfEntitlement ,
                "NoDaysTakenApproved": str(TakenApproved) + " (" + str(DaysOfApproved) + " + "+ str(DaysOfPending) + ") ",
                "NoDaysLeft": DaysOfleft,
                "type_id": rec["type"],
                "type": list(filter(lambda r: (r["leave_type_id"].upper() == rec["type"]), leaveTypeLst))[0].get("leave_type"),
                "calendarYear": getDisplayLeaveYear(rec["year"]),
                "officeHeader": str_officeHeader,
                "submit_date": rec['submit_date'],
                "details": leaveDetailsLst
                }
            
            #Output to array to excel file
            approvalRecordLst.append(leaveRecord)

    # Get mapping from MongoDB
    rpt = reportMap.find_one ( { "report": "Application Form"} )

    #filename when using in Heroku:
    fs = gridfs.GridFS(db)
    wb = load_workbook(filename=BytesIO(fs.get(ObjectId(rpt["file"]["fileObj"])).read()))
    ws = wb[rpt["file"]["wsName"]]

    try:
        genApplyForm(ws, approvalRecordLst, rpt)
    except Exception as e:
        print (e)
        return jsonify({"error_message" : "Sorry, we failed to generate Application form"}), 501    

    # Output 
    out = BytesIO()
    wb.save(out)
    out.seek(0)

    wb.close()            
    print('sending file...')

    return send_file(out,  attachment_filename='a_file.xls', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

         

# Status_code 200: passed
# Status_code 501: failed, Not enough days left for the leave
# Status_code 502: failed, Overlapped leave period
# Status_code 503: failed, Leave Type Not Found
# Status_code 504: failed, Staff Record Not Exist
# Status_code 505: failed, Incorrect parameters
# Status_code 506: failed, Over 14 days

@eleave.route("/api/applyleave", methods=['POST'])
@checkLogged.check_logged
#@app.route("/api/applyleave", methods=['POST'])
def apiApplyLeave():
    psInput = request.get_json()
    result = applyLeave(psInput)
    try: 
        return jsonify(result), result['Status_code'] # APP
    except:
        return jsonify(result) # postman

@eleave.route("/api/listapprove", methods=['POST'])
@checkLogged.check_logged
#@app.route("/api/listapprove", methods=['POST'])
def apiListApprove():    
    psInput = request.get_json()
    result = listApprove(psInput)
    try: 
        return jsonify(result), result['Status_code'] # APP
    except:
        return jsonify(result) # postman

@eleave.route("/api/listleave", methods=['POST'])
@checkLogged.check_logged
#@app.route("/api/listleave", methods=['POST'])
#@app.route("/")
def apiListLeave():
    psInput = request.get_json()
    result = listLeave(psInput)
    try: 
        return jsonify(result), result['Status_code'] # APP
    except:
        return jsonify(result) # postman

@eleave.route("/api/ListApprovedByYear", methods=['POST'])
@checkLogged.check_logged
def apiListApprovedByYear():    
    psInput = request.get_json()
    result = listApprovedLeaveByYear(psInput)
    try: 
        return jsonify(result), result['Status_code'] # APP
    except:
        return jsonify(result) # postman

# Status_code 200: passed
# Status_code 505: failed, Incorrect parameters
# Status_code 504: failed, Staff Record Not Exist
# Status_code 603: failed, Current Status cannot be changed
# Status_code 604: failed, Only applicant himself / herself can cancel leave
# Status_code 605: failed, Leave cancel already submitted and waiting for approval
# Status_code 606: failed, Cannot cancel leave in the past period
# Status_code 607: failed, User is not the first approver
# Status_code 608: failed, User is not the second approver
# Status_code 609: failed, User is not the third approver
# Status_code 610: failed, Leave already approved
# Status_code 611: failed, Incorrect action

@eleave.route("/api/changestatus", methods=['POST'])
@checkLogged.check_logged
#@app.route("/api/changestatus", methods=['POST'])
def apiChangeStatus():
    psInput = request.get_json()
    result = changeStatus(psInput)
    try: 
        return jsonify(result), result['Status_code'] # APP
    except:
        return jsonify(result) # postman
 
